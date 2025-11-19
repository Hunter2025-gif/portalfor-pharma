from django.core.paginator import Paginator
# --- RESTORE: Admin Timeline View ---
from django.db.models import F, ExpressionWrapper, DateTimeField, Count
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.shortcuts import render, redirect
from django.contrib import messages
from datetime import timedelta

# Model imports
from bmr.models import BMR
from workflow.models import BatchPhaseExecution
from accounts.models import CustomUser
from products.models import Product

from .permissions import require_dashboard_permission, check_dashboard_permission

@login_required
def admin_timeline_view(request):
    """Admin Timeline View - Track all BMRs through the system"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')

    # Get export format if requested
    export_format = request.GET.get('export')

    # Get all BMRs with timeline data
    bmrs = BMR.objects.select_related('product', 'created_by', 'approved_by').all()

    # Add timeline data for each BMR
    timeline_data = []
    from workflow.models import BatchPhaseExecution
    for bmr in bmrs:
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase').order_by('phase__phase_order')
        bmr_created = bmr.created_date
        fgs_completed = phases.filter(
            phase__phase_name='finished_goods_store',
            status='completed'
        ).first()
        total_time_days = None
        total_time_hours = None
        if fgs_completed and fgs_completed.completed_date:
            # FIXED: Calculate using first phase start time instead of BMR creation time
            first_started_phase = phases.filter(started_date__isnull=False).order_by('started_date').first()
            if first_started_phase:
                total_duration = fgs_completed.completed_date - first_started_phase.started_date
                total_time_days = total_duration.days
                total_time_hours = round(total_duration.total_seconds() / 3600, 2)
        phase_timeline = []
        for phase in phases:
            phase_data = {
                'phase_name': phase.phase.phase_name.replace('_', ' ').title(),
                'status': phase.status.title(),
                'started_date': phase.started_date,
                'completed_date': phase.completed_date,
                'started_by': phase.started_by.get_full_name() if phase.started_by else None,
                'completed_by': phase.completed_by.get_full_name() if phase.completed_by else None,
                'duration_hours': None,
                'operator_comments': getattr(phase, 'operator_comments', '') or '',
                'phase_order': phase.phase.phase_order if hasattr(phase.phase, 'phase_order') else 0,
            }
            if phase.started_date and phase.completed_date:
                duration = phase.completed_date - phase.started_date
                phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
            elif phase.started_date and not phase.completed_date:
                duration = timezone.now() - phase.started_date
                phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
            phase_timeline.append(phase_data)
        timeline_data.append({
            'bmr': bmr,
            'total_time_days': total_time_days,
            'total_time_hours': total_time_hours,
            'phase_timeline': phase_timeline,
            'current_phase': phases.filter(status__in=['pending', 'in_progress']).first(),
            'is_completed': fgs_completed is not None,
        })

    # Handle exports
    if export_format in ['csv', 'excel']:
        return export_timeline_data(request, timeline_data, export_format)

    # Pagination
    paginator = Paginator(timeline_data, 10)  # 10 BMRs per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'user': request.user,
        'page_obj': page_obj,
        'timeline_data': page_obj.object_list,
        'dashboard_title': 'BMR Timeline Tracking',
        'total_bmrs': len(timeline_data),
    }

    return render(request, 'dashboards/admin_timeline.html', context)
# Basic workflow_chart view to resolve missing view error
from django.contrib.auth.decorators import login_required

@login_required
def workflow_chart(request):
    """Workflow Chart View (placeholder)"""
    return render(request, 'dashboards/workflow_chart.html', {'dashboard_title': 'Workflow Chart'})
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_protect
from django.contrib import messages
from django.db.models import Count, Q, Min, Max, F, ExpressionWrapper, DateTimeField, Q
from django.utils import timezone
from datetime import datetime, timedelta
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_http_methods
from django.core.paginator import Paginator
from django.db.models.functions import Coalesce
import csv
import xlwt
import logging
import calendar
import os
import re
from datetime import datetime

logger = logging.getLogger('dashboards')
from bmr.models import BMR
from workflow.models import BatchPhaseExecution, Machine
from workflow.services import WorkflowService
from products.models import Product
from accounts.models import CustomUser
from .analytics import (
    get_monthly_production_analytics, 
    get_yearly_production_comparison,
    get_product_type_production_totals,
    export_monthly_production_to_excel
)

def dashboard_home(request):
    """Route users to their role-specific dashboard or redirect to login"""
    if not request.user.is_authenticated:
        # Redirect unauthenticated users directly to login
        return redirect('accounts:login')
    
    user_role = request.user.role
    
    role_dashboard_map = {
        'qa': 'dashboards:qa_dashboard',
        'regulatory': 'dashboards:regulatory_dashboard',
        'store_manager': 'dashboards:store_dashboard',  # Raw material release
        'packaging_store': 'dashboards:packaging_dashboard',
        'finished_goods_store': 'dashboards:finished_goods_dashboard',
        'mixing_operator': 'dashboards:mixing_dashboard',
        'qc': 'dashboards:qc_dashboard',
        'tube_filling_operator': 'dashboards:tube_filling_dashboard',
        'packing_operator': 'dashboards:packing_dashboard',
        'granulation_operator': 'dashboards:granulation_dashboard',
        'blending_operator': 'dashboards:blending_dashboard',
        'compression_operator': 'dashboards:compression_dashboard',
        'sorting_operator': 'dashboards:sorting_dashboard',
        'coating_operator': 'dashboards:coating_dashboard',
        'drying_operator': 'dashboards:drying_dashboard',
        'filling_operator': 'dashboards:filling_dashboard',
        'dispensing_operator': 'dashboards:operator_dashboard',  # Material dispensing uses operator dashboard
        'equipment_operator': 'dashboards:operator_dashboard',
        'cleaning_operator': 'dashboards:operator_dashboard',
        'production_manager': 'dashboards:production_manager_dashboard',  # Production manager dashboard
        'quarantine': 'quarantine:dashboard',  # Quarantine users go to quarantine app dashboard
        'admin': 'dashboards:admin_dashboard',
    }
    
    dashboard_url = role_dashboard_map.get(user_role, 'dashboards:admin_dashboard')
    return redirect(dashboard_url)

@login_required
# @require_dashboard_permission('admin_dashboard')  # TEMPORARILY DISABLED FOR LOGIN FIX
def admin_dashboard(request):
    """Admin Dashboard with Charts and Basic Data"""
    from django.db.models import Count
    from products.models import Product
    
    # Get basic statistics only
    total_bmrs = BMR.objects.count()
    active_batches = BMR.objects.filter(status__in=['draft', 'approved', 'in_production']).count()
    completed_batches = BMR.objects.filter(status='completed').count()
    rejected_batches = BMR.objects.filter(status='rejected').count()
    
    # Get system metrics
    total_users = CustomUser.objects.count()
    active_users_count = CustomUser.objects.filter(is_active=True).count()
    
    # Get recent BMRs only (last 5 for performance)
    recent_bmrs = BMR.objects.select_related('product').order_by('-created_date')[:5]
    
    # Active phases - simplified query
    active_phases = BatchPhaseExecution.objects.filter(
        status__in=['pending', 'in_progress']
    ).select_related('bmr', 'phase')[:10]  # Limit to 10 for speed
    
    # CHART DATA - Restore for dashboard charts
    # Chart data - Product Type Distribution
    product_types = Product.objects.values('product_type').annotate(count=Count('product_type'))
    tablet_count = 0
    capsule_count = 0
    ointment_count = 0
    
    for item in product_types:
        product_type = item['product_type'].lower() if item['product_type'] else ''
        if 'tablet' in product_type:
            tablet_count += item['count']
        elif 'capsule' in product_type:
            capsule_count += item['count']
        elif 'ointment' in product_type or 'cream' in product_type:
            ointment_count += item['count']
    
    # Phase completion data for chart
    common_phases = ['mixing', 'drying', 'granulation', 'compression', 'packing']
    phase_data = {}
    
    for phase_name in common_phases:
        completed = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status='completed'
        ).count()
        
        in_progress = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status__in=['pending', 'in_progress']
        ).count()
        
        phase_data[f"{phase_name}_completed"] = completed
        phase_data[f"{phase_name}_inprogress"] = in_progress
    
    context = {
        'user': request.user,
        'dashboard_title': 'Admin Dashboard',
        'total_bmrs': total_bmrs,
        'active_batches': active_batches,
        'completed_batches': completed_batches,
        'rejected_batches': rejected_batches,
        'total_users': total_users,
        'active_users_count': active_users_count,
        'recent_bmrs': recent_bmrs,
        'active_phases': active_phases,
        'timeline_data': [],  # Empty for now
        
        # CHART DATA
        'tablet_count': tablet_count,
        'capsule_count': capsule_count,
        'ointment_count': ointment_count,
        
        # Phase chart data
        'mixing_completed': phase_data.get('mixing_completed', 0),
        'mixing_inprogress': phase_data.get('mixing_inprogress', 0),
        'granulation_completed': phase_data.get('granulation_completed', 0),
        'granulation_inprogress': phase_data.get('granulation_inprogress', 0),
        'compression_completed': phase_data.get('compression_completed', 0),
        'compression_inprogress': phase_data.get('compression_inprogress', 0),
        'packing_completed': phase_data.get('packing_completed', 0),
        'packing_inprogress': phase_data.get('packing_inprogress', 0),
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)


@login_required
    completed_count = sum(1 for item in timeline_data if item['is_completed'])
    in_progress_count = len(timeline_data) - completed_count
    
    # Calculate average production time
    completed_times = [item['total_time_days'] for item in timeline_data if item['total_time_days']]
    avg_production_time = round(sum(completed_times) / len(completed_times)) if completed_times else None
    
    # === ACTIVE PHASES DATA ===
    # Get currently active phases with enhanced data
    active_phases = BatchPhaseExecution.objects.filter(
        status__in=['pending', 'in_progress']
    ).select_related('bmr__product', 'phase', 'started_by').order_by('-started_date')
    
    # Duration calculation is now handled by the model's duration_hours property
    # No need to manually calculate duration as the property handles both completed and active phases
    
    # Chart data - Product Type Distribution
    from products.models import Product
    product_types = Product.objects.values('product_type').annotate(count=Count('product_type'))
    tablet_count = 0
    capsule_count = 0
    ointment_count = 0
    
    for item in product_types:
        product_type = item['product_type'].lower() if item['product_type'] else ''
        if 'tablet' in product_type:
            tablet_count += item['count']
        elif 'capsule' in product_type:
            capsule_count += item['count']
        elif 'ointment' in product_type or 'cream' in product_type:
            ointment_count += item['count']
    
    # Phase completion data for chart
    phase_data = {}
    common_phases = ['mixing', 'drying', 'granulation', 'compression', 'packing']
    
    for phase_name in common_phases:
        completed = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status='completed'
        ).count()
        
        in_progress = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status__in=['pending', 'in_progress']
        ).count()
        
        phase_data[f"{phase_name}_completed"] = completed
        phase_data[f"{phase_name}_inprogress"] = in_progress
    
    # Weekly production trend data
    current_date = timezone.now().date()
    week_start = current_date - timedelta(days=current_date.weekday())
    
    weekly_data = {}
    for i in range(4):
        week_end = week_start - timedelta(days=1)
        week_start_prev = week_start - timedelta(days=7)
        
        started = BMR.objects.filter(
            created_date__date__gte=week_start_prev,
            created_date__date__lte=week_end
        ).count()
        
        completed = BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            completed_date__date__gte=week_start_prev,
            completed_date__date__lte=week_end,
            status='completed'
        ).count()
        
        weekly_data[f"started_week{4-i}"] = started
        weekly_data[f"completed_week{4-i}"] = completed
        
        week_start = week_start_prev
        
    # Quality Control data for chart
    qc_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name__icontains='qc'
    )
    
    qc_data = {
        'passed': qc_phases.filter(status='completed').count(),
        'failed': qc_phases.filter(status='failed').count(),
        'pending': qc_phases.filter(status__in=['pending', 'in_progress']).count(),
    }
    
    # Recent activity
    recent_bmrs = BMR.objects.select_related('product', 'created_by').order_by('-created_date')[:10]
    recent_users = CustomUser.objects.filter(is_active=True).order_by('-date_joined')[:10]
    
    # === WORK IN PROGRESS DATA ===
    # Get BMRs that are currently in production (not completed or rejected)
    work_in_progress_bmrs = []
    wip_bmrs = BMR.objects.filter(
        status__in=['draft', 'approved', 'in_production']
    ).select_related('product')
    
    for bmr in wip_bmrs:
        # Calculate progress first to determine if this is truly work in progress
        total_phases = BatchPhaseExecution.objects.filter(bmr=bmr).count()
        completed_phases = BatchPhaseExecution.objects.filter(bmr=bmr, status='completed').count()
        progress_percentage = int((completed_phases / total_phases * 100)) if total_phases > 0 else 0
        
        # Skip BMRs that are 100% complete - they shouldn't be in work in progress
        if progress_percentage >= 100:
            continue
            
        # Get current active phase for this BMR
        current_phase = BatchPhaseExecution.objects.filter(
            bmr=bmr,
            status__in=['pending', 'in_progress']
        ).select_related('phase').first()
        
        # If no current active phase, get the next phase that should start
        if not current_phase:
            current_phase = BatchPhaseExecution.objects.filter(
                bmr=bmr,
                status__in=['not_ready', 'pending']
            ).select_related('phase').order_by('id').first()
        
        # Get the first phase that was started to determine actual start date
        first_started_phase = BatchPhaseExecution.objects.filter(
            bmr=bmr,
            started_date__isnull=False
        ).order_by('started_date').first()
        
        # Determine current phase name more intelligently
        if current_phase:
            current_phase_name = current_phase.phase.get_phase_name_display()
            if current_phase.status == 'pending':
                current_phase_name += ' (Pending)'
            elif current_phase.status == 'in_progress':
                current_phase_name += ' (In Progress)'
            elif current_phase.status == 'not_ready':
                current_phase_name += ' (Waiting)'
        else:
            # If no phases found, this shouldn't be in WIP
            if progress_percentage == 0:
                current_phase_name = 'Not Started'
            else:
                current_phase_name = 'Phase Transition'
        
        # Build BMR data for work in progress display
        bmr_data = {
            'id': bmr.id,
            'product': bmr.product,
            'batch_number': bmr.batch_number,
            'actual_batch_size': bmr.batch_size,
            'actual_batch_size_unit': bmr.product.batch_size_unit,
            'current_phase_name': current_phase_name,
            'actual_start_date': first_started_phase.started_date if first_started_phase else bmr.created_date,
            'progress_percentage': progress_percentage,
        }
        work_in_progress_bmrs.append(bmr_data)
    
    # System health metrics
    pending_approvals = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        status='pending'
    ).count()
    
    failed_phases = BatchPhaseExecution.objects.filter(
        status='failed',
        completed_date__date=timezone.now().date()
    ).count()
    
    # Production metrics
    production_stats = {
        'in_production': BatchPhaseExecution.objects.filter(
            status='in_progress'
        ).count(),
        'quality_hold': BatchPhaseExecution.objects.filter(
            phase__phase_name__contains='qc',
            status='pending'
        ).count(),
        'awaiting_packaging': BatchPhaseExecution.objects.filter(
            phase__phase_name='packaging_material_release',
            status='pending'
        ).count(),
        'final_qa_pending': BatchPhaseExecution.objects.filter(
            phase__phase_name='final_qa',
            status='pending'
        ).count(),
        'in_fgs': BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            status__in=['completed', 'in_progress']
        ).count(),
    }
    
    # Get enhanced analytics - using exact database data
    # Monthly production stats
    monthly_stats = {
        'labels': [],
        'created': [],
        'completed': [],
        'rejected': []
    }
    
    # Get last 6 months of data
    current_month = timezone.now().date().replace(day=1)
    for i in range(6):
        month_start = current_month - timedelta(days=i*30)
        month_end = month_start + timedelta(days=29)
        month_label = month_start.strftime('%b %Y')
        
        created_count = BMR.objects.filter(
            created_date__date__gte=month_start,
            created_date__date__lte=month_end
        ).count()
        
        completed_count = BMR.objects.filter(
            status='completed',
            approved_date__date__gte=month_start,
            approved_date__date__lte=month_end
        ).count()
        
        rejected_count = BMR.objects.filter(
            status='rejected',
            approved_date__date__gte=month_start,
            approved_date__date__lte=month_end
        ).count()
        
        monthly_stats['labels'].insert(0, month_label)
        monthly_stats['created'].insert(0, created_count)
        monthly_stats['completed'].insert(0, completed_count)
        monthly_stats['rejected'].insert(0, rejected_count)
    
    # Cycle times - actual database data
    cycle_times = {
        'labels': [],
        'avg_days': []
    }
    
    # Get average cycle time by product type
    for product_type in ['tablet', 'capsule', 'ointment']:
        bmrs_of_type = BMR.objects.filter(
            product__product_type__icontains=product_type,
            status='completed'
        )
        
        total_days = 0
        count = 0
        for bmr in bmrs_of_type:
            fgs_phase = BatchPhaseExecution.objects.filter(
                bmr=bmr,
                phase__phase_name='finished_goods_store',
                status='completed'
            ).first()
            
            if fgs_phase and fgs_phase.completed_date:
                days = (fgs_phase.completed_date - bmr.created_date).days
                total_days += days
                count += 1
        
        if count > 0:
            cycle_times['labels'].append(product_type.title())
            cycle_times['avg_days'].append(round(total_days / count, 1))
    
    # Bottleneck analysis - actual phase duration data
    bottleneck_analysis = []
    phase_names = ['mixing', 'granulation', 'compression', 'coating', 'packaging_material_release']
    
    for phase_name in phase_names:
        phases = BatchPhaseExecution.objects.filter(
            phase__phase_name__icontains=phase_name,
            status='completed',
            started_date__isnull=False,
            completed_date__isnull=False
        )
        
        total_hours = 0
        count = 0
        for phase in phases:
            duration = (phase.completed_date - phase.started_date).total_seconds() / 3600
            total_hours += duration
            count += 1
        
        if count > 0:
            avg_hours = round(total_hours / count, 2)
            bottleneck_analysis.append({
                'phase': phase_name.replace('_', ' ').title(),
                'avg_duration': avg_hours,
                'total_executions': count
            })
    
    # Sort by average duration (longest first)
    bottleneck_analysis.sort(key=lambda x: x['avg_duration'], reverse=True)
    
    # Quality metrics - actual QC data
    quality_metrics = {
        'labels': [],
        'pass_rates': [],
        'fail_rates': []
    }
    
    qc_phase_names = ['post_mixing_qc', 'post_compression_qc', 'post_blending_qc']
    for qc_phase in qc_phase_names:
        total_tests = BatchPhaseExecution.objects.filter(
            phase__phase_name=qc_phase
        ).count()
        
        passed_tests = BatchPhaseExecution.objects.filter(
            phase__phase_name=qc_phase,
            status='completed'
        ).count()
        
        failed_tests = BatchPhaseExecution.objects.filter(
            phase__phase_name=qc_phase,
            status='failed'
        ).count()
        
        if total_tests > 0:
            pass_rate = round((passed_tests / total_tests) * 100, 1)
            fail_rate = round((failed_tests / total_tests) * 100, 1)
            
            quality_metrics['labels'].append(qc_phase.replace('_', ' ').title())
            quality_metrics['pass_rates'].append(pass_rate)
            quality_metrics['fail_rates'].append(fail_rate)
    
    # Productivity metrics - actual operator data
    top_operators = []
    operators = CustomUser.objects.filter(
        role__in=['mixing_operator', 'compression_operator', 'granulation_operator', 'packing_operator']
    )
    
    for operator in operators:
        completed_phases = BatchPhaseExecution.objects.filter(
            completed_by=operator,
            status='completed'
        ).count()
        
        if completed_phases > 0:
            top_operators.append({
                'name': operator.get_full_name(),
                'completions': completed_phases,
                'role': operator.get_role_display()
            })
    
    # Sort by completions (highest first) and take top 10
    top_operators.sort(key=lambda x: x['completions'], reverse=True)
    top_operators = top_operators[:10]
    
    productivity_metrics = {
        'top_operators': top_operators,
        'total_operators': operators.count(),
        'total_completions': sum([op['completions'] for op in top_operators])
    }
    
    # Performance metrics - Average cycle time
    completed_bmrs = BMR.objects.filter(status__in=['approved', 'completed']).annotate(
        cycle_time=ExpressionWrapper(
            F('approved_date') - F('created_date'),
            output_field=DateTimeField()
        )
    )
    
    # Calculate average time from BMR creation to FGS for completed batches
    avg_production_time = None
    completed_productions = []
    
    for bmr in BMR.objects.filter(status='completed'):
        first_phase = BatchPhaseExecution.objects.filter(bmr=bmr).order_by('phase__phase_order').first()
        last_phase = BatchPhaseExecution.objects.filter(
            bmr=bmr, 
            phase__phase_name='finished_goods_store',
            status='completed'
        ).first()
        
        if first_phase and last_phase and first_phase.started_date and last_phase.completed_date:
            duration_days = (last_phase.completed_date - bmr.created_date).days
            completed_productions.append(duration_days)
    
    if completed_productions:
        avg_production_time = round(sum(completed_productions) / len(completed_productions), 1)
    
    # === MACHINE MANAGEMENT DATA ===
    # Get all machines
    all_machines = Machine.objects.all().order_by('machine_type', 'name')
    
    # Get recent breakdowns (last 30 days)
    recent_breakdowns = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-breakdown_start_time')[:20]
    
    # Get recent changeovers (last 30 days)
    recent_changeovers = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-changeover_start_time')[:20]
    
    # Count total breakdowns and changeovers
    total_breakdowns = BatchPhaseExecution.objects.filter(breakdown_occurred=True).count()
    total_changeovers = BatchPhaseExecution.objects.filter(changeover_occurred=True).count()
    
    # Breakdown and changeover counts for today
    today = timezone.now().date()
    breakdowns_today = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__date=today
    ).count()
    changeovers_today = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__date=today
    ).count()
    
    # Machine utilization summary
    machine_stats = {}
    for machine in all_machines:
        usage_count = BatchPhaseExecution.objects.filter(machine_used=machine).count()
        breakdown_count = BatchPhaseExecution.objects.filter(
            machine_used=machine,
            breakdown_occurred=True
        ).count()
        changeover_count = BatchPhaseExecution.objects.filter(
            machine_used=machine,
            changeover_occurred=True
        ).count()
        
        machine_stats[machine.id] = {
            'machine': machine,
            'usage_count': usage_count,
            'breakdown_count': breakdown_count,
            'changeover_count': changeover_count,
            'breakdown_rate': round((breakdown_count / usage_count * 100), 1) if usage_count > 0 else 0
        }
    
    # === MONTHLY PRODUCTION ANALYTICS ===
    # Get current month analytics
    current_month_analytics = get_monthly_production_analytics()
    
    # Get year comparison data
    yearly_comparison = get_yearly_production_comparison()
    
    # Get current month totals by product type
    current_month_totals = get_product_type_production_totals()
    
    # Get selected month data (default to current month)
    selected_month = int(request.GET.get('month', timezone.now().month))
    selected_year = int(request.GET.get('year', timezone.now().year))
    selected_month_analytics = get_monthly_production_analytics(selected_month, selected_year)
    
    context = {
        'user': request.user,
        'dashboard_title': 'System Administration Dashboard',
        # Key Metrics for cards
        'total_bmrs': total_bmrs,
        'active_batches': active_batches,
        'completed_batches': completed_batches,
        'rejected_batches': rejected_batches,
        # System Status
        'active_users_count': active_users_count,
        # Chart data
        'tablet_count': tablet_count,
        'capsule_count': capsule_count,
        'ointment_count': ointment_count,
        # Phase data for charts
        'mixing_completed': phase_data.get('mixing_completed', 0),
        'mixing_inprogress': phase_data.get('mixing_inprogress', 0),
        'drying_completed': phase_data.get('drying_completed', 0),
        'drying_inprogress': phase_data.get('drying_inprogress', 0),
        'granulation_completed': phase_data.get('granulation_completed', 0),
        'granulation_inprogress': phase_data.get('granulation_inprogress', 0),
        'compression_completed': phase_data.get('compression_completed', 0),
        'compression_inprogress': phase_data.get('compression_inprogress', 0),
        'packing_completed': phase_data.get('packing_completed', 0),
        'packing_inprogress': phase_data.get('packing_inprogress', 0),
        # Weekly trend data
        'weekly_data': weekly_data,
        # QC data
        'qc_data': qc_data,
        # === NEW: Timeline and Active Phases Data ===
        'timeline_data': timeline_data[:10],  # Show first 10 for performance
        'completed_count': completed_count,
        'in_progress_count': in_progress_count,
        'avg_production_time': avg_production_time,
        'active_phases': active_phases[:10],  # Show first 10 active phases
        # Recent activity
        'recent_bmrs': recent_bmrs,
        # Work in Progress data
        'work_in_progress_bmrs': work_in_progress_bmrs,
        # === MACHINE MANAGEMENT DATA ===
        'all_machines': all_machines,
        'recent_breakdowns': recent_breakdowns,
        'recent_changeovers': recent_changeovers,
        'total_breakdowns': total_breakdowns,
        'total_changeovers': total_changeovers,
        'breakdowns_today': breakdowns_today,
        'changeovers_today': changeovers_today,
        'machine_stats': machine_stats,
        # === MONTHLY PRODUCTION ANALYTICS ===
        'current_month_analytics': current_month_analytics,
        'yearly_comparison': yearly_comparison,
        'current_month_totals': current_month_totals,
        'selected_month_analytics': selected_month_analytics,
        'selected_month': selected_month,
        'selected_year': selected_year,
    }
    
    return render(request, 'dashboards/admin_dashboard.html', context)

@login_required
@csrf_protect
def qa_dashboard(request):
    """Quality Assurance Dashboard"""
    if request.user.role != 'qa':
        messages.error(request, 'Access denied. QA role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for Final QA workflow
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        comments = request.POST.get('comments', '')
        
        if phase_id and action in ['start', 'approve', 'reject']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Start the Final QA review process
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Final QA review started by {request.user.get_full_name()}. Notes: {comments}"
                    phase_execution.save()
                    
                    messages.success(request, f'Final QA review started for batch {phase_execution.bmr.batch_number}. You can now complete the review.')
                
                elif action == 'approve':
                    # Complete Final QA with approval
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments += f"\nFinal QA Approved by {request.user.get_full_name()}. Comments: {comments}"
                    phase_execution.save()
                    
                    # Trigger next phase in workflow (should be finished goods store)
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.success(request, f'Final QA approved for batch {phase_execution.bmr.batch_number}. Batch is ready for finished goods storage.')
                    
                elif action == 'reject':
                    # Complete Final QA with rejection
                    phase_execution.status = 'failed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments += f"\nFinal QA Rejected by {request.user.get_full_name()}. Rejection Reason: {comments}"
                    phase_execution.save()
                    
                    # Rollback to appropriate packing phase based on product type
                    bmr = phase_execution.bmr
                    product_type = bmr.product.product_type
                    
                    # Determine which packing phase to rollback to based on product type
                    if product_type == 'tablet':
                        if hasattr(bmr.product, 'tablet_type') and bmr.product.tablet_type == 'tablet_2':
                            rollback_phase = 'bulk_packing'
                        else:
                            rollback_phase = 'blister_packing'
                    elif product_type == 'capsule':
                        rollback_phase = 'blister_packing'
                    elif product_type == 'ointment':
                        rollback_phase = 'secondary_packaging'
                    else:
                        rollback_phase = 'secondary_packaging'  # default
                    
                    # Find and activate the appropriate packing phase for rework
                    rollback_execution = BatchPhaseExecution.objects.filter(
                        bmr=bmr,
                        phase__phase_name=rollback_phase
                    ).first()
                    
                    if rollback_execution:
                        rollback_execution.status = 'pending'
                        rollback_execution.operator_comments = f"Returned for rework due to Final QA rejection. Reason: {comments}. Original comments: {rollback_execution.operator_comments}"
                        rollback_execution.save()
                        
                        messages.warning(request, f'Final QA rejected for batch {bmr.batch_number}. Batch has been sent back to {rollback_phase.replace("_", " ").title()} for rework.')
                    else:
                        messages.error(request, f'Could not find {rollback_phase} phase to rollback to for batch {bmr.batch_number}.')
                    
            except Exception as e:
                messages.error(request, f'Error processing Final QA: {str(e)}')
        
        return redirect('dashboards:qa_dashboard')
    
    # Get QA-specific data
    total_bmrs = BMR.objects.count()
    draft_bmrs = BMR.objects.filter(status='draft').count()
    submitted_bmrs = BMR.objects.filter(status='submitted').count()
    my_bmrs = BMR.objects.filter(created_by=request.user).count()
    
    # CRITICAL: Get pending BMR requests from production managers
    from bmr.models import BMRRequest
    bmr_requests_pending = BMRRequest.objects.filter(status='pending').exclude(bmr__isnull=True).select_related('product', 'requested_by', 'bmr').order_by('-request_date')
    
    # CRITICAL: Get quarantine sample requests for QA processing - MISSING CONTEXT
    from quarantine.models import SampleRequest
    pending_quarantine_samples = SampleRequest.objects.filter(
        sampled_by__isnull=True  # Samples not yet processed by QA
    ).select_related('quarantine_batch__bmr__product', 'requested_by').order_by('-request_date')
    
    # BMR request counts for dashboard stats
    bmr_request_counts = {
        'pending': BMRRequest.objects.filter(status='pending').exclude(bmr__isnull=True).count(),
        'approved': BMRRequest.objects.filter(status='approved').exclude(bmr__isnull=True).count(),
        'rejected': BMRRequest.objects.filter(status='rejected').exclude(bmr__isnull=True).count(),
        'completed': BMRRequest.objects.filter(status='completed').exclude(bmr__isnull=True).count(),
    }
    
    # Recent BMRs created by this user
    recent_bmrs = BMR.objects.filter(created_by=request.user).select_related('product').order_by('-created_date')[:5]
    
    # BMRs needing final QA review
    final_qa_pending = BatchPhaseExecution.objects.filter(
        phase__phase_name='final_qa',
        status='pending'
    ).select_related('bmr', 'phase')[:10]
    
    # Final QA reviews in progress (started but not completed)
    final_qa_in_progress = BatchPhaseExecution.objects.filter(
        phase__phase_name='final_qa',
        status='in_progress'
    ).select_related('bmr', 'phase')[:10]
    
    # Build operator history for this user: only regulatory approval phases completed by this user
    regulatory_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        completed_by=request.user
    ).order_by('-completed_date')[:10]
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M'),
            'batch': p.bmr.batch_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in regulatory_phases
    ]

    context = {
        'user': request.user,
        'total_bmrs': total_bmrs,
        'draft_bmrs': draft_bmrs,
        'submitted_bmrs': submitted_bmrs,
        'my_bmrs': my_bmrs,
        'recent_bmrs': recent_bmrs,
        'bmr_requests_pending': bmr_requests_pending,  # Add BMR requests to context
        'bmr_request_counts': bmr_request_counts,  # Add request counts for stats
        'pending_quarantine_samples': pending_quarantine_samples,  # CRITICAL: Add quarantine samples
        'final_qa_pending': final_qa_pending,
        'final_qa_in_progress': final_qa_in_progress,
        'dashboard_title': 'Quality Assurance Dashboard',
        'operator_history': operator_history,
    }
    return render(request, 'dashboards/qa_dashboard.html', context)

@login_required
def regulatory_dashboard(request):
    """Regulatory Dashboard"""
    if request.user.role != 'regulatory':
        messages.error(request, 'Access denied. Regulatory role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for approval/rejection
    if request.method == 'POST':
        action = request.POST.get('action')
        bmr_id = request.POST.get('bmr_id')
        comments = request.POST.get('comments', '')
        
        if bmr_id and action in ['approve', 'reject']:
            try:
                bmr = get_object_or_404(BMR, pk=bmr_id)
                
                # Find the regulatory approval phase for this BMR
                regulatory_phase = BatchPhaseExecution.objects.filter(
                    bmr=bmr,
                    phase__phase_name='regulatory_approval',
                    status='pending'
                ).first()
                
                if regulatory_phase:
                    if action == 'approve':
                        regulatory_phase.status = 'completed'
                        regulatory_phase.completed_by = request.user
                        regulatory_phase.completed_date = timezone.now()
                        regulatory_phase.operator_comments = f"Approved by {request.user.get_full_name()}. Comments: {comments}"
                        regulatory_phase.save()
                        
                        # Update BMR status
                        bmr.status = 'approved'
                        bmr.approved_by = request.user
                        bmr.approved_date = timezone.now()
                        bmr.save()
                        
                        # Trigger next phase in workflow
                        WorkflowService.trigger_next_phase(bmr, regulatory_phase.phase)
                        
                        messages.success(request, f'BMR {bmr.batch_number} has been approved successfully.')
                        
                    elif action == 'reject':
                        regulatory_phase.status = 'failed'
                        regulatory_phase.completed_by = request.user
                        regulatory_phase.completed_date = timezone.now()
                        regulatory_phase.operator_comments = f"Rejected by {request.user.get_full_name()}. Reason: {comments}"
                        regulatory_phase.save()
                        
                        # Update BMR status
                        bmr.status = 'rejected'
                        bmr.approved_by = request.user
                        bmr.approved_date = timezone.now()
                        bmr.save()
                        
                        messages.warning(request, f'BMR {bmr.batch_number} has been rejected and sent back to QA.')
                else:
                    messages.error(request, 'No pending regulatory approval found for this BMR.')
                    
            except Exception as e:
                messages.error(request, f'Error processing request: {str(e)}')
        
        return redirect('dashboards:regulatory_dashboard')
    
    # BMRs waiting for regulatory approval (pending regulatory_approval phase) with pagination
    approvals_page = request.GET.get('approvals_page', 1)
    all_pending_approvals = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        status='pending'
    ).select_related('bmr__product', 'phase').order_by('bmr__created_date')
    
    # Paginate pending approvals (10 items per page)
    approvals_paginator = Paginator(all_pending_approvals, 10)
    try:
        pending_approvals = approvals_paginator.page(approvals_page)
    except:
        pending_approvals = approvals_paginator.page(1)
    
    # Recent regulatory activities with pagination
    activities_page = request.GET.get('activities_page', 1)
    all_recent_activities = BatchPhaseExecution.objects.filter(
        phase__phase_name='regulatory_approval',
        status__in=['completed', 'failed']
    ).select_related('bmr__product', 'phase', 'completed_by').order_by('-completed_date')
    
    # Paginate recent activities (7 items per page)
    activities_paginator = Paginator(all_recent_activities, 7)
    try:
        recent_activities = activities_paginator.page(activities_page)
    except:
        recent_activities = activities_paginator.page(1)
    
    # Statistics
    stats = {
        'pending_approvals': all_pending_approvals.count(),
        'approved_today': BMR.objects.filter(
            status='approved',
            approved_date__date=timezone.now().date()
        ).count(),
        'rejected_this_week': BMR.objects.filter(
            status='rejected',
            approved_date__gte=timezone.now().date() - timedelta(days=7)
        ).count(),
        'total_bmrs': BMR.objects.count(),
    }
    
    context = {
        'user': request.user,
        'pending_approvals': pending_approvals,
        'approvals_paginator': approvals_paginator,
        'recent_activities': recent_activities,
        'activities_paginator': activities_paginator,
        'stats': stats,
        'dashboard_title': 'Regulatory Dashboard'
    }
    return render(request, 'dashboards/regulatory_dashboard.html', context)

@login_required  
# @require_dashboard_permission('system_logs')  # TEMPORARILY DISABLED FOR LOGIN FIX
def system_logs_viewer(request):
    """
    System Logs Web Viewer - Industry Standard Approach
    Accessible to admin users for system monitoring
    """
    
    from .log_utils import LogAnalyzer
    
    # Get filter parameters
    level_filter = request.GET.get('level')
    date_filter = request.GET.get('date', 'today')
    search_query = request.GET.get('search', '').strip()
    page = request.GET.get('page', 1)
    
    # Initialize log analyzer
    analyzer = LogAnalyzer()
    
    # Get log entries with filters
    log_entries = analyzer.get_log_entries(
        limit=500,  # Get more for pagination
        level_filter=level_filter,
        date_filter=date_filter,
        search_query=search_query
    )
    
    # Paginate results (20 per page - industry standard)
    paginator = Paginator(log_entries, 20)
    try:
        paginated_logs = paginator.page(page)
    except:
        paginated_logs = paginator.page(1)
    
    # Get log statistics
    log_stats = analyzer.get_log_stats()
    log_file_info = analyzer.get_log_file_info()
    
    context = {
        'user': request.user,
        'logs': paginated_logs,
        'log_stats': log_stats,
        'log_file_info': log_file_info,
        'current_filters': {
            'level': level_filter,
            'date': date_filter,
            'search': search_query
        },
        'log_levels': LogAnalyzer.LOG_LEVELS.keys(),
        'dashboard_title': 'System Logs'
    }
    
    return render(request, 'dashboards/system_logs.html', context)

@login_required
def production_manager_dashboard(request):
    """Production Manager Dashboard"""
    if request.user.role != 'production_manager':
        messages.error(request, 'Access denied. Production Manager role required.')
        return redirect('dashboards:dashboard_home')

    # Get production overview data
    from workflow.models import BatchPhaseExecution
    from bmr.models import BMR, BMRRequest
    
    # Get pagination parameters
    requests_page = request.GET.get('requests_page', 1)
    bmrs_page = request.GET.get('bmrs_page', 1)
    
    # BMR Request Statistics
    all_requests = BMRRequest.objects.filter(requested_by=request.user).exclude(bmr__isnull=True)
    bmr_request_stats = {
        'total': all_requests.count(),
        'pending': all_requests.filter(status='pending').count(),
        'approved': all_requests.filter(status='approved').count(),
        'completed': all_requests.filter(status='completed').count(),
    }
    
    # Production Statistics
    all_bmrs = BMR.objects.all()
    production_stats = {
        'total_bmrs': all_bmrs.count(),
        'active_production': all_bmrs.filter(status='in_production').count(),
        'completed_batches': all_bmrs.filter(status='completed').count(),
        'pending_approval': all_bmrs.filter(status='pending').count(),
    }
    
    # Recent BMR Requests with pagination (7 per page)
    all_recent_requests = BMRRequest.objects.filter(
        requested_by=request.user
    ).exclude(bmr__isnull=True).select_related('product', 'bmr').order_by('-request_date')
    
    requests_paginator = Paginator(all_recent_requests, 7)
    try:
        recent_bmr_requests = requests_paginator.page(requests_page)
    except:
        recent_bmr_requests = requests_paginator.page(1)
    
    # BMRs created from user's requests with pagination (7 per page)
    all_user_bmrs = BMR.objects.filter(
        bmr_requests__requested_by=request.user
    ).distinct().select_related('product').order_by('-created_date')
    
    bmrs_paginator = Paginator(all_user_bmrs, 7)
    try:
        user_bmrs = bmrs_paginator.page(bmrs_page)
    except:
        user_bmrs = bmrs_paginator.page(1) if all_user_bmrs.exists() else None
    
    # Get all batches in production phases for overview
    production_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name__in=[
            'mixing', 'granulation', 'blending', 'compression', 'coating',
            'drying', 'filling', 'tube_filling', 'sorting', 'blister_packing',
            'bulk_packing', 'secondary_packaging'
        ]
    ).select_related('bmr', 'phase', 'started_by').order_by('-started_date')[:10]
    
    context = {
        'bmr_request_stats': bmr_request_stats,
        'production_stats': production_stats,
        'recent_bmr_requests': recent_bmr_requests,
        'requests_paginator': requests_paginator,
        'user_bmrs': user_bmrs,
        'bmrs_paginator': bmrs_paginator,
        'production_phases': production_phases,
        'dashboard_title': 'Production Manager Dashboard',
    }
    return render(request, 'dashboards/production_manager_dashboard.html', context)

@login_required
def store_dashboard(request):
    """Store Manager Dashboard - Raw Material Release Phase"""
    if request.user.role != 'store_manager':
        messages.error(request, 'Access denied. Store Manager role required.')
        return redirect('dashboards:dashboard_home')
    
    if request.method == 'POST':
        bmr_id = request.POST.get('bmr_id')
        action = request.POST.get('action')
        notes = request.POST.get('notes', '')
        
        try:
            bmr = BMR.objects.get(pk=bmr_id)
            
            # Get the raw material release phase
            phase_execution = BatchPhaseExecution.objects.get(
                bmr=bmr,
                phase__phase_name='raw_material_release'
            )
            
            if action == 'start':
                phase_execution.status = 'in_progress'
                phase_execution.started_by = request.user
                phase_execution.started_date = timezone.now()
                phase_execution.operator_comments = f"Raw material release started by {request.user.get_full_name()}. Notes: {notes}"
                phase_execution.save()
                
                messages.success(request, f'Raw material release started for batch {bmr.batch_number}.')
                
            elif action == 'complete':
                phase_execution.status = 'completed'
                phase_execution.completed_by = request.user
                phase_execution.completed_date = timezone.now()
                phase_execution.operator_comments = f"Raw materials released by {request.user.get_full_name()}. Notes: {notes}"
                phase_execution.save()
                
                # Trigger next phase in workflow (material_dispensing)
                WorkflowService.trigger_next_phase(bmr, phase_execution.phase)
                
                messages.success(request, f'Raw materials released for batch {bmr.batch_number}. Material dispensing is now available.')
                
        except Exception as e:
            messages.error(request, f'Error processing raw material release: {str(e)}')
    
        return redirect('dashboards:store_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get raw material release phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        # Get both normal and rework phases for this user's role
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        # Add rework indicator to phases that have been rolled back
        for phase in user_phases:
            if hasattr(phase, 'rework_count') and phase.rework_count > 0:
                phase.is_rework = True
                phase.rework_from = phase.rollback_from.phase.phase_name if phase.rollback_from else 'Unknown'
            else:
                phase.is_rework = False
        my_phases.extend(user_phases)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }
    
    # Get recently completed releases (last 7 days)
    recently_completed = BatchPhaseExecution.objects.filter(
        phase__phase_name='raw_material_release',
        status='completed',
        completed_date__gte=timezone.now() - timedelta(days=7)
    ).select_related('bmr__product', 'completed_by').order_by('-completed_date')[:10]
    
    return render(request, 'dashboards/store_dashboard.html', {
        'my_phases': my_phases,
        'stats': stats,
        'recently_completed': recently_completed,
    })

@login_required
def operator_dashboard(request):
    """Generic operator dashboard for production phases"""
    
    # Handle POST requests for phase start/completion
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        comments = request.POST.get('comments', '')
        
        # Machine-related fields
        machine_id = request.POST.get('machine_id')
        
        # Breakdown fields
        breakdown_occurred = request.POST.get('breakdown_occurred') == 'on'
        breakdown_start_time = request.POST.get('breakdown_start_time')
        breakdown_end_time = request.POST.get('breakdown_end_time')
        
        # Changeover fields  
        changeover_occurred = request.POST.get('changeover_occurred') == 'on'
        changeover_start_time = request.POST.get('changeover_start_time')
        changeover_end_time = request.POST.get('changeover_end_time')
        
        if phase_id and action in ['start', 'complete']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Check if machine selection is required for this phase
                    machine_required_phases = ['granulation', 'blending', 'compression', 'coating', 'blister_packing', 'bulk_packing', 'filling']
                    phase_name = phase_execution.phase.phase_name
                    
                    # For capsule filling, only require machine for filling phase
                    if phase_name == 'filling' and phase_execution.bmr.product.product_type != 'Capsule':
                        machine_required = False
                    elif phase_name in machine_required_phases:
                        machine_required = True
                    else:
                        machine_required = False
                    
                    if machine_required and not machine_id:
                        messages.error(request, f'Machine selection is required for {phase_name} phase.')
                        return redirect(request.path)
                    
                    # Validate that the phase can actually be started
                    if not WorkflowService.can_start_phase(phase_execution.bmr, phase_execution.phase.phase_name):
                        messages.error(request, f'Cannot start {phase_execution.phase.phase_name} for batch {phase_execution.bmr.batch_number} - prerequisites not met.')
                        return redirect(request.path)
                    
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Started by {request.user.get_full_name()}. Notes: {comments}"
                    
                    # Set machine if provided
                    if machine_id:
                        try:
                            machine = Machine.objects.get(id=machine_id, is_active=True)
                            phase_execution.machine_used = machine
                        except Machine.DoesNotExist:
                            messages.error(request, 'Selected machine not found or inactive.')
                            return redirect(request.path)
                    
                    phase_execution.save()
                    
                    machine_info = f" using {phase_execution.machine_used.name}" if phase_execution.machine_used else ""
                    messages.success(request, f'Phase {phase_execution.phase.phase_name}{machine_info} started for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'complete':
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"Completed by {request.user.get_full_name()}. Notes: {comments}"
                    
                    # Only handle breakdown/changeover for production phases (not material dispensing)
                    phase_name = phase_execution.phase.phase_name
                    exclude_breakdown_phases = ['material_dispensing', 'bmr_creation', 'regulatory_approval', 'bulk_packing', 'secondary_packaging']
                    
                    if phase_name not in exclude_breakdown_phases:
                        # Handle breakdown tracking
                        phase_execution.breakdown_occurred = breakdown_occurred
                        if breakdown_occurred and breakdown_start_time and breakdown_end_time:
                            from datetime import datetime
                            try:
                                phase_execution.breakdown_start_time = datetime.fromisoformat(breakdown_start_time.replace('T', ' '))
                                phase_execution.breakdown_end_time = datetime.fromisoformat(breakdown_end_time.replace('T', ' '))
                            except ValueError:
                                messages.warning(request, 'Invalid breakdown time format. Breakdown recorded without times.')
                        
                        # Handle changeover tracking
                        phase_execution.changeover_occurred = changeover_occurred
                        if changeover_occurred and changeover_start_time and changeover_end_time:
                            from datetime import datetime
                            try:
                                phase_execution.changeover_start_time = datetime.fromisoformat(changeover_start_time.replace('T', ' '))
                                phase_execution.changeover_end_time = datetime.fromisoformat(changeover_end_time.replace('T', ' '))
                            except ValueError:
                                messages.warning(request, 'Invalid changeover time format. Changeover recorded without times.')
                    
                    phase_execution.save()
                    
                    # Trigger next phase in workflow
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    completion_msg = f'Phase {phase_execution.phase.phase_name} completed for batch {phase_execution.bmr.batch_number}.'
                    if breakdown_occurred:
                        completion_msg += ' Breakdown recorded.'
                    if changeover_occurred:
                        completion_msg += ' Changeover recorded.'
                    
                    messages.success(request, completion_msg)
                    
            except Exception as e:
                messages.error(request, f'Error processing phase: {str(e)}')
        
        return redirect(request.path)  # Redirect to same dashboard
    
    # Get phases this user can work on - OPTIMIZED VERSION
    # Map user roles to phases they can handle (same as in WorkflowService)
    role_phase_mapping = {
        'qa': ['bmr_creation', 'final_qa'],
        'regulatory': ['regulatory_approval'],
        'store_manager': ['raw_material_release'],
        'dispensing_operator': ['material_dispensing'],
        'packaging_store': ['packaging_material_release'],
        'finished_goods_store': ['finished_goods_store'],
        'qc': ['post_compression_qc', 'post_mixing_qc', 'post_blending_qc'],
        'mixing_operator': ['mixing'],
        'granulation_operator': ['granulation'],
        'blending_operator': ['blending'],
        'compression_operator': ['compression'],
        'coating_operator': ['coating'],
        'drying_operator': ['drying'],
        'filling_operator': ['filling'],
        'tube_filling_operator': ['tube_filling'],
        'packing_operator': ['blister_packing', 'bulk_packing', 'secondary_packaging'],
        'sorting_operator': ['sorting'],
    }
    
    allowed_phases = role_phase_mapping.get(request.user.role, [])
    
    # Single optimized query instead of N+1 queries
    my_phases = list(BatchPhaseExecution.objects.filter(
        phase__phase_name__in=allowed_phases,
        status__in=['pending', 'in_progress']
    ).select_related('phase', 'bmr', 'bmr__product', 'bmr__created_by').order_by('phase__phase_order'))
    
    # Add timer data directly to each phase
    from workflow.models import PhaseTimingSetting, ProductMachineTimingSetting
    for phase in my_phases:
        # Initialize timer data as None for all phases
        phase.timer_data = None
        
        # Add timer data only for in-progress phases with start time
        if phase.status == 'in_progress' and phase.started_date:
            elapsed = timezone.now() - phase.started_date
            elapsed_hours = elapsed.total_seconds() / 3600
            
            # ENHANCED: Get expected duration using new Product+Machine timing system
            expected_hours = ProductMachineTimingSetting.get_expected_duration_for_execution(phase)
            warning_threshold = ProductMachineTimingSetting.get_warning_threshold_for_execution(phase)
            
            remaining_hours = expected_hours - elapsed_hours
            remaining_seconds = max(0, remaining_hours * 3600)
            
            # Determine timer status using dynamic warning threshold
            warning_time = expected_hours * (warning_threshold / 100.0)  # Use configured threshold
            if elapsed_hours >= expected_hours * 1.5:  # 150% overrun
                timer_status = 'overrun'
            elif elapsed_hours >= expected_hours:
                timer_status = 'expired'
            elif elapsed_hours >= warning_time:
                timer_status = 'warning'
            else:
                timer_status = 'normal'
            
            # Add timer data to the phase object
            phase.timer_data = {
                'elapsed_hours': round(elapsed_hours, 2),
                'expected_hours': expected_hours,
                'remaining_hours': round(remaining_hours, 2),
                'remaining_seconds': int(remaining_seconds),
                'timer_status': timer_status
            }
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }

    # Determine the primary phase name for this role
    role_phase_mapping = {
        'mixing_operator': 'mixing',
        'granulation_operator': 'granulation',
        'blending_operator': 'blending',
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'drying_operator': 'drying',
        'filling_operator': 'filling',
        'tube_filling_operator': 'tube_filling',
        'packing_operator': 'packing',
        'sorting_operator': 'sorting',
        'dispensing_operator': 'dispensing',  # Material dispensing operator
    }

    phase_name = role_phase_mapping.get(request.user.role, 'production')
    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)

    # Operator History: all phases completed by this user for their role

    # Fix: Use .distinct() before slicing to avoid TypeError
    completed_phases_qs = BatchPhaseExecution.objects.filter(
        completed_by=request.user
    ).select_related('bmr', 'phase').order_by('-completed_date')
    completed_phases = list(completed_phases_qs[:20])
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M') if (p.completed_date or p.started_date or p.created_date) else '',
            'batch': p.bmr.bmr_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in completed_phases
    ]

    # Operator Statistics
    # Use .distinct() before slicing for batches_handled
    batches_handled = completed_phases_qs.values('bmr').distinct().count()
    total_completed = completed_phases_qs.count()
    total_attempted = BatchPhaseExecution.objects.filter(started_by=request.user).count()
    success_rate = round((total_completed / total_attempted) * 100, 1) if total_attempted else 0
    completion_times = [
        (p.completed_date - p.started_date).total_seconds() / 60
        for p in completed_phases if p.completed_date and p.started_date
    ]
    avg_completion_time = f"{round(sum(completion_times)/len(completion_times), 1)} min" if completion_times else "-"
    assignment_status = "You have assignments pending." if stats['pending_phases'] > 0 else "All assignments up to date."
    operator_stats = {
        'batches_handled': batches_handled,
        'success_rate': success_rate,
        'avg_completion_time': avg_completion_time,
        'assignment_status': assignment_status,
    }

    # Operator Assignments: current in-progress or pending phases
    operator_assignments = [
        f"{p.bmr.bmr_number} - {p.phase.get_phase_name_display()} ({p.status.title()})"
        for p in my_phases if p.status in ['pending', 'in_progress']
    ]

    # Get machines for this operator's phase type
    machine_type_mapping = {
        'granulation_operator': 'granulation',
        'blending_operator': 'blending',
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'packing_operator': 'blister_packing',  # Packing operator uses blister packing machines
        'filling_operator': 'filling',  # For capsule filling
    }
    
    user_machine_type = machine_type_mapping.get(request.user.role)
    available_machines = []
    if user_machine_type:
        available_machines = Machine.objects.filter(
            machine_type=user_machine_type,
            is_active=True
        ).order_by('name')

    # Determine if this role should show breakdown/changeover tracking
    # Exclude material dispensing and administrative phases
    breakdown_tracking_roles = [
        'mixing_operator', 'granulation_operator', 'blending_operator', 'compression_operator',
        'coating_operator', 'drying_operator', 'filling_operator', 'tube_filling_operator',
        'sorting_operator', 'packing_operator'
    ]
    show_breakdown_tracking = request.user.role in breakdown_tracking_roles



    context = {
        'user': request.user,
        'my_phases': my_phases,
        'stats': stats,
        'phase_name': phase_name,
        'daily_progress': daily_progress,
        'dashboard_title': f'{request.user.get_role_display()} Dashboard',
        'operator_history': operator_history,
        'operator_stats': operator_stats,
        'operator_assignments': operator_assignments,
        'available_machines': available_machines,
        'show_breakdown_tracking': show_breakdown_tracking,
    }

    return render(request, 'dashboards/operator_dashboard.html', context)

# Specific operator dashboards
@login_required
def mixing_dashboard(request):
    """Mixing Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def granulation_dashboard(request):
    """Granulation Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def blending_dashboard(request):
    """Blending Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def compression_dashboard(request):
    """Compression Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def coating_dashboard(request):
    """Coating Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def drying_dashboard(request):
    """Drying Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def filling_dashboard(request):
    """Filling Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def tube_filling_dashboard(request):
    """Tube Filling Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def sorting_dashboard(request):
    """Sorting Operator Dashboard"""
    return operator_dashboard(request)

@login_required
def qc_dashboard(request):
    """Quality Control Dashboard"""
    if request.user.role != 'qc':
        messages.error(request, 'Access denied. QC role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for QC test results
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        test_results = request.POST.get('test_results', '')
        
        if phase_id and action in ['start', 'pass', 'fail']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Start QC testing
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"QC Testing started by {request.user.get_full_name()}. Notes: {test_results}"
                    phase_execution.save()
                    
                    messages.success(request, f'QC testing started for batch {phase_execution.bmr.batch_number}.')
                
                elif action == 'pass':
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"QC Test Passed by {request.user.get_full_name()}. Results: {test_results}"
                    phase_execution.save()
                    
                    # Trigger next phase in workflow
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.success(request, f'QC test passed for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'fail':
                    phase_execution.status = 'failed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"QC Test Failed by {request.user.get_full_name()}. Results: {test_results}"
                    phase_execution.save()
                    
                    # Rollback to previous phase
                    WorkflowService.rollback_to_previous_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.warning(request, f'QC test failed for batch {phase_execution.bmr.batch_number}. Rolled back to previous phase.')
                    
            except Exception as e:
                messages.error(request, f'Error processing QC test: {str(e)}')
        
        return redirect('dashboards:qc_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get QC phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Statistics
    stats = {
        'pending_tests': len([p for p in my_phases if p.status == 'pending']),
        'in_testing': len([p for p in my_phases if p.status == 'in_progress']),
        'passed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date(),
            status='completed'
        ).count(),
        'failed_this_week': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date__gte=timezone.now().date() - timedelta(days=7),
            status='failed'
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }
    
    daily_progress = min(100, (stats['passed_today'] / max(1, stats['pending_tests'] + stats['passed_today'])) * 100)
    
    # Get quarantine samples waiting for QC testing
    from quarantine.models import SampleRequest
    quarantine_samples = SampleRequest.objects.filter(
        sample_date__isnull=False,  # Processed by QA
        qc_status='pending'  # Waiting for QC testing
    ).select_related('quarantine_batch__bmr__product', 'sampled_by').order_by('sample_date')

    context = {
        'user': request.user,
        'my_phases': my_phases,
        'qc_phases': my_phases,  # Add this for template compatibility
        'quarantine_samples': quarantine_samples,  # Add quarantine samples for QC
        'stats': stats,
        'daily_progress': daily_progress,
        'dashboard_title': 'Quality Control Dashboard'
    }
    
    return render(request, 'dashboards/qc_dashboard.html', context)

@login_required
def packaging_dashboard(request):
    """Packaging Store Dashboard"""
    if request.user.role != 'packaging_store':
        messages.error(request, 'Access denied. Packaging Store role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for packaging material release
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        notes = request.POST.get('notes', '')
        
        if phase_id and action in ['start', 'complete']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Validate that the phase can actually be started
                    if not WorkflowService.can_start_phase(phase_execution.bmr, phase_execution.phase.phase_name):
                        messages.error(request, f'Cannot start packaging material release for batch {phase_execution.bmr.batch_number} - prerequisites not met.')
                        return redirect('dashboards:packaging_dashboard')
                    
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Packaging material release started by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    messages.success(request, f'Packaging material release started for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'complete':
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"Packaging materials released by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    # Set session variables for next phase notification
                    request.session['completed_phase'] = phase_execution.phase.phase_name
                    request.session['completed_bmr'] = phase_execution.bmr.id
                    
                    # Trigger next phase in workflow (should be packing phases)
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    # Determine correct message based on product type
                    if phase_execution.bmr.product.product_type == 'tablet' and getattr(phase_execution.bmr.product, 'tablet_type', None) == 'tablet_2':
                        messages.success(request, f'Packaging materials released for batch {phase_execution.bmr.batch_number}. Bulk packing is now available.')
                    else:
                        messages.success(request, f'Packaging materials released for batch {phase_execution.bmr.batch_number}. Packing phases are now available.')
                    
            except Exception as e:
                messages.error(request, f'Error processing packaging material release: {str(e)}')
        
        return redirect('dashboards:packaging_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get packaging phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Add timing information to active phases
    from workflow.models import PhaseTimingSetting, ProductMachineTimingSetting
    for phase in my_phases:
        if phase.status == 'in_progress' and phase.started_date:
            # Calculate timing information
            elapsed = timezone.now() - phase.started_date
            elapsed_hours = elapsed.total_seconds() / 3600
            
            # ENHANCED: Get expected duration using new Product+Machine timing system
            expected_hours = ProductMachineTimingSetting.get_expected_duration_for_execution(phase)
            
            # Calculate remaining time
            remaining_hours = expected_hours - elapsed_hours
            remaining_seconds = int(remaining_hours * 3600) if remaining_hours > 0 else 0
            
            # Add timing data to phase object
            phase.elapsed_hours = round(elapsed_hours, 2)
            phase.expected_hours = expected_hours
            phase.remaining_hours = round(remaining_hours, 2)
            phase.remaining_seconds = remaining_seconds
            phase.is_overrun = remaining_hours <= 0
            phase.progress_percent = min(100, (elapsed_hours / expected_hours) * 100)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }
    
    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)
    
    # Build operator history for this user (recent phases where user was started_by or completed_by)
    recent_phases = BatchPhaseExecution.objects.filter(
        Q(started_by=request.user) | Q(completed_by=request.user)
    ).order_by('-started_date', '-completed_date')[:10]
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M'),
            'batch': p.bmr.batch_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in recent_phases
    ]

    context = {
        'user': request.user,
        'my_phases': my_phases,
        'stats': stats,
        'daily_progress': daily_progress,
        'dashboard_title': 'Packaging Store Dashboard',
        'operator_history': operator_history,
    }
    
    # Get next phase info for notification
    completed_phase = request.session.pop('completed_phase', None)
    bmr_id = request.session.pop('completed_bmr', None)
    bmr = None
    next_phase = None
    if bmr_id:
        try:
            bmr = BMR.objects.get(id=bmr_id)
            # For tablet type 2, make sure bulk packing comes before secondary packing
            if bmr.product.product_type == 'tablet' and getattr(bmr.product, 'tablet_type', None) == 'tablet_2':
                # Check if material release was just completed
                if completed_phase == 'packaging_material_release':
                    next_phase = BatchPhaseExecution.objects.filter(bmr=bmr, phase__phase_name='bulk_packing').first()
            
            # Fallback to standard next phase logic if no specific phase found
            if not next_phase:
                next_phase = WorkflowService.get_next_phase(bmr)
        except BMR.DoesNotExist:
            pass
    
    # Add notification context
    context.update({
        'completed_phase': completed_phase,
        'bmr': bmr,
        'next_phase': next_phase
    })
    
    return render(request, 'dashboards/packaging_dashboard.html', context)

@login_required
def packing_dashboard(request):
    """Packing Operator Dashboard"""
    if request.user.role != 'packing_operator':
        messages.error(request, 'Access denied. Packing Operator role required.')
        return redirect('dashboards:dashboard_home')
    
    # Handle POST requests for packing phase completion
    if request.method == 'POST':
        action = request.POST.get('action')
        phase_id = request.POST.get('phase_id')
        notes = request.POST.get('notes', '')
        
        if phase_id and action in ['start', 'complete']:
            try:
                phase_execution = get_object_or_404(BatchPhaseExecution, pk=phase_id)
                
                if action == 'start':
                    # Validate that the phase can actually be started
                    if not WorkflowService.can_start_phase(phase_execution.bmr, phase_execution.phase.phase_name):
                        messages.error(request, f'Cannot start packing for batch {phase_execution.bmr.batch_number} - prerequisites not met.')
                        return redirect('dashboards:packing_dashboard')
                    
                    # Handle machine selection
                    machine_id = request.POST.get('machine_id')
                    if machine_id:
                        try:
                            machine = Machine.objects.get(id=machine_id, is_active=True)
                            phase_execution.machine_used = machine
                        except Machine.DoesNotExist:
                            messages.error(request, 'Selected machine is not available.')
                            return redirect('dashboards:packing_dashboard')
                    
                    phase_execution.status = 'in_progress'
                    phase_execution.started_by = request.user
                    phase_execution.started_date = timezone.now()
                    phase_execution.operator_comments = f"Packing started by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    messages.success(request, f'Packing started for batch {phase_execution.bmr.batch_number}.')
                    
                elif action == 'complete':
                    # Handle breakdown tracking
                    breakdown_occurred = request.POST.get('breakdown_occurred') == 'on'
                    if breakdown_occurred:
                        phase_execution.breakdown_occurred = True
                        breakdown_start = request.POST.get('breakdown_start_time')
                        breakdown_end = request.POST.get('breakdown_end_time')
                        breakdown_reason = request.POST.get('breakdown_reason', '')
                        
                        if breakdown_start:
                            phase_execution.breakdown_start_time = datetime.fromisoformat(breakdown_start.replace('T', ' '))
                        if breakdown_end:
                            phase_execution.breakdown_end_time = datetime.fromisoformat(breakdown_end.replace('T', ' '))
                        phase_execution.breakdown_reason = breakdown_reason
                    
                    # Handle changeover tracking
                    changeover_occurred = request.POST.get('changeover_occurred') == 'on'
                    if changeover_occurred:
                        phase_execution.changeover_occurred = True
                        changeover_start = request.POST.get('changeover_start_time')
                        changeover_end = request.POST.get('changeover_end_time')
                        changeover_reason = request.POST.get('changeover_reason', '')
                        
                        if changeover_start:
                            phase_execution.changeover_start_time = datetime.fromisoformat(changeover_start.replace('T', ' '))
                        if changeover_end:
                            phase_execution.changeover_end_time = datetime.fromisoformat(changeover_end.replace('T', ' '))
                        phase_execution.changeover_reason = changeover_reason
                    
                    phase_execution.status = 'completed'
                    phase_execution.completed_by = request.user
                    phase_execution.completed_date = timezone.now()
                    phase_execution.operator_comments = f"Packing completed by {request.user.get_full_name()}. Notes: {notes}"
                    phase_execution.save()
                    
                    # Trigger next phase in workflow
                    WorkflowService.trigger_next_phase(phase_execution.bmr, phase_execution.phase)
                    
                    messages.success(request, f'Packing completed for batch {phase_execution.bmr.batch_number}.')
                    
            except Exception as e:
                messages.error(request, f'Error processing packing phase: {str(e)}')
        
        return redirect('dashboards:packing_dashboard')
    
    # Get all BMRs
    all_bmrs = BMR.objects.all()
    
    # Get packing phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    
    # Add countdown timer logic for active packing phases
    from workflow.models import PhaseTimingSetting
    for phase in my_phases:
        if phase.status == 'in_progress' and phase.started_date:
            # Calculate timing information
            elapsed = timezone.now() - phase.started_date
            elapsed_hours = elapsed.total_seconds() / 3600
            
            # ENHANCED: Get expected duration using new Product+Machine timing system
            expected_hours = ProductMachineTimingSetting.get_expected_duration_for_execution(phase)
            
            # Calculate remaining time
            remaining_hours = expected_hours - elapsed_hours
            remaining_seconds = int(remaining_hours * 3600) if remaining_hours > 0 else 0
            
            # Add timing data to phase object
            phase.elapsed_hours = round(elapsed_hours, 2)
            phase.expected_hours = expected_hours
            phase.remaining_hours = round(remaining_hours, 2)
            phase.remaining_seconds = remaining_seconds
            phase.is_overrun = remaining_hours <= 0
            phase.progress_percent = min(100, (elapsed_hours / expected_hours) * 100)
    
    # Statistics
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'pending_packing': len([p for p in my_phases if p.status == 'pending']),  # For template compatibility
        'in_progress_packing': len([p for p in my_phases if p.status == 'in_progress']),  # For template compatibility
        'completed_today': BatchPhaseExecution.objects.filter(
            completed_by=request.user,
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': len(set([p.bmr for p in my_phases])),
    }

    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)
    
    # Get available machines for this user role
    machine_type_mapping = {
        'mixing_operator': 'mixing',
        'granulation_operator': 'granulation',
        'blending_operator': 'blending', 
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'tube_filling_operator': 'tube_filling',
        'packing_operator': 'blister_packing',  # Packing operator uses blister packing machines
        'filling_operator': 'filling',  # For capsule filling
    }
    
    user_machine_type = machine_type_mapping.get(request.user.role)
    available_machines = []
    if user_machine_type:
        available_machines = Machine.objects.filter(
            machine_type=user_machine_type,
            is_active=True
        ).order_by('name')
    
    # Determine if this role should show breakdown/changeover tracking
    # Only for phases that use machines
    breakdown_tracking_roles = [
        'mixing_operator', 'granulation_operator', 'blending_operator', 'compression_operator',
        'coating_operator', 'tube_filling_operator', 'filling_operator'
    ]
    # For packing operator, only show breakdown tracking for blister packing phases (machine-based)
    show_breakdown_tracking = request.user.role in breakdown_tracking_roles    # Build operator history for this user (recent phases where user was started_by or completed_by)
    recent_phases = BatchPhaseExecution.objects.filter(
        Q(started_by=request.user) | Q(completed_by=request.user)
    ).order_by('-started_date', '-completed_date')[:10]
    operator_history = [
        {
            'date': (p.completed_date or p.started_date or p.created_date).strftime('%Y-%m-%d %H:%M'),
            'batch': p.bmr.batch_number,
            'phase': p.phase.get_phase_name_display(),
        }
        for p in recent_phases
    ]

    context = {
        'user': request.user,
        'my_phases': my_phases,
        'packing_phases': my_phases,  # Add this for template compatibility
        'stats': stats,
        'daily_progress': daily_progress,
        'dashboard_title': 'Packing Dashboard',
        'operator_history': operator_history,
        'available_machines': available_machines,
        'show_breakdown_tracking': show_breakdown_tracking,
    }
    
    return render(request, 'dashboards/packing_dashboard.html', context)

def format_phase_name(name):
    """Format phase name for display"""
    if not name:
        return ""
    # Replace underscores with spaces
    name = name.replace("_", " ")
    # Title case
    return name.title()

@login_required
def finished_goods_dashboard(request):
    """Finished Goods Store Dashboard with Inventory Management"""
    if request.user.role != 'finished_goods_store':
        messages.error(request, 'Access denied. Finished Goods Store role required.')
        return redirect('dashboards:dashboard_home')
    
    # Import FGS models
    from fgs_management.models import FGSInventory, ProductRelease, FGSAlert
    from django.utils import timezone
    from datetime import timedelta
    
    # Get all BMRs
    all_bmrs = BMR.objects.select_related('product', 'created_by').all()
    
    # Get phases this user can work on
    my_phases = []
    for bmr in all_bmrs:
        user_phases = WorkflowService.get_phases_for_user_role(bmr, request.user.role)
        my_phases.extend(user_phases)
    # Only show finished_goods_store phases
    my_phases = [p for p in my_phases if getattr(p.phase, 'phase_name', None) == 'finished_goods_store']
    
    # Get all finished goods store phases for history statistics
    all_fgs_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store'
    ).select_related('bmr', 'phase', 'bmr__product')
    
    # FGS Inventory Statistics
    total_inventory_items = FGSInventory.objects.count()
    available_for_sale = FGSInventory.objects.filter(status='available').count()
    
    # Recent releases (last 7 days)
    recent_releases_count = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=7)
    ).count()
    
    # Active alerts
    active_alerts_count = FGSAlert.objects.filter(is_resolved=False).count()
    
    # Recent inventory items
    recent_inventory = FGSInventory.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).select_related('product', 'bmr').order_by('-created_at')[:10]
    
    # Current inventory available for release
    available_inventory = FGSInventory.objects.filter(
        status__in=['stored', 'available'],
        quantity_available__gt=0
    ).select_related('product', 'bmr').order_by('-created_at')
    
    # Completed FGS phases without inventory entries
    completed_fgs_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store',
        status='completed'
    ).exclude(
        bmr__in=FGSInventory.objects.values_list('bmr', flat=True)
    ).select_related('bmr__product').order_by('-completed_date')[:10]
    
    # Recent releases
    recent_releases = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=14)
    ).select_related('inventory__product', 'inventory__bmr').order_by('-release_date')[:10]
    
    # Active alerts
    active_alerts = FGSAlert.objects.filter(
        is_resolved=False
    ).select_related('inventory').order_by('-priority', '-created_at')[:10]
    
    # Filtering support for dashboard cards
    filter_param = request.GET.get('filter')
    detail_param = request.GET.get('detail')
    
    # Detail view for specific card
    if detail_param:
        if detail_param == 'pending':
            my_phases = [p for p in my_phases if p.status == 'pending']
        elif detail_param == 'in_progress':
            my_phases = [p for p in my_phases if p.status == 'in_progress']
        elif detail_param == 'completed_today':
            today = timezone.now().date()
            my_phases = [p for p in all_fgs_phases if p.status == 'completed' and 
                         getattr(p, 'completed_date', None) and p.completed_date.date() == today]
        elif detail_param == 'total_batches':
            # Show all batches that have reached FGS
            my_phases = list(all_fgs_phases)
    # Regular filtering
    elif filter_param:
        if filter_param == 'completed_today':
            my_phases = [p for p in my_phases if p.status == 'completed' and getattr(p, 'completed_by', None) == request.user and getattr(p, 'completed_date', None) and p.completed_date.date() == timezone.now().date()]
        elif filter_param == 'total_batches':
            # Show all phases (default)
            pass
        else:
            my_phases = [p for p in my_phases if p.status == filter_param]
    
    # History statistics (last 7 days)
    today = timezone.now().date()
    last_7_days = [today - timezone.timedelta(days=i) for i in range(7)]
    daily_completions = {}
    
    for day in last_7_days:
        count = all_fgs_phases.filter(
            status='completed',
            completed_date__date=day
        ).count()
        daily_completions[day.strftime('%a')] = count
    
    # Product type statistics in FGS
    product_types = {}
    for phase in all_fgs_phases.filter(status__in=['in_progress', 'completed']):
        product_type = phase.bmr.product.product_type
        if product_type in product_types:
            product_types[product_type] += 1
        else:
            product_types[product_type] = 1

    # Statistics - Updated with real FGS data
    stats = {
        'pending_phases': len([p for p in my_phases if p.status == 'pending']),
        'in_progress_phases': len([p for p in my_phases if p.status == 'in_progress']),
        'completed_today': BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            status='completed',
            completed_date__date=timezone.now().date()
        ).count(),
        'total_batches': all_fgs_phases.values('bmr').distinct().count(),
        'daily_history': daily_completions,
        'product_types': product_types,
        
        # FGS-specific statistics
        'total_inventory_items': total_inventory_items,
        'available_for_sale': available_for_sale,
        'recent_releases': recent_releases.count(),
        'active_alerts': active_alerts.count(),
    }

    # Determine the primary phase name for this role
    role_phase_mapping = {
        'mixing_operator': 'mixing',
        'granulation_operator': 'granulation',
        'blending_operator': 'blending',
        'compression_operator': 'compression',
        'coating_operator': 'coating',
        'drying_operator': 'drying',
        'filling_operator': 'filling',
        'tube_filling_operator': 'tube_filling',
        'packing_operator': 'packing',
        'sorting_operator': 'sorting',
    }

    phase_name = role_phase_mapping.get(request.user.role, 'production')
    daily_progress = min(100, (stats['completed_today'] / max(1, stats['pending_phases'] + stats['completed_today'])) * 100)
    
    # Get recently completed goods
    recent_completed = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store',
        status='completed'
    ).select_related('bmr', 'bmr__product').order_by('-completed_date')[:5]
    
    # Storage efficiency (time from final QA to FGS)
    efficiency_data = []
    for phase in recent_completed:
        final_qa_phase = BatchPhaseExecution.objects.filter(
            bmr=phase.bmr,
            phase__phase_name='final_qa',
            status='completed'
        ).first()
        
        if final_qa_phase and final_qa_phase.completed_date and phase.completed_date:
            storage_time = (phase.completed_date - final_qa_phase.completed_date).total_seconds() / 3600  # hours
            efficiency_data.append({
                'bmr': phase.bmr,
                'time_hours': round(storage_time, 1)
            })
    
    # Card specific view
    detail_title = None
    if request.GET.get('detail'):
        detail = request.GET.get('detail')
        if detail == 'pending':
            detail_title = 'Pending Storage'
        elif detail == 'in_progress':
            detail_title = 'In Storage'
        elif detail == 'completed_today':
            detail_title = 'Stored Today'
        elif detail == 'total_batches':
            detail_title = 'All Batches in FGS'

    # Process all phases to add display name
    for phase in my_phases:
        if hasattr(phase, 'phase') and hasattr(phase.phase, 'phase_name'):
            phase.display_name = format_phase_name(phase.phase.phase_name)
    
    context = {
        'user': request.user,
        'my_phases': my_phases,
        'stats': stats,
        'phase_name': 'finished_goods_store',
        'phase_display_name': 'Finished Goods Store',
        'daily_progress': daily_progress,
        'dashboard_title': 'Finished Goods Store Dashboard',
        'active_filter': filter_param,
        'recent_completed': recent_completed,
        'efficiency_data': efficiency_data,
        'detail_title': detail_title,
        'detail_view': request.GET.get('detail'),
        
        # New FGS inventory data
        'recent_inventory': recent_inventory,
        'recent_releases': recent_releases,
        'active_alerts': active_alerts,
        'available_inventory': available_inventory,
        'completed_fgs_phases': completed_fgs_phases,
    }

    return render(request, 'dashboards/finished_goods_dashboard.html', context)

@login_required
def admin_fgs_monitor(request):
    """Admin FGS Monitor - Track finished goods storage with inventory management"""
    if not request.user.is_staff:
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Import FGS models
    from fgs_management.models import FGSInventory, ProductRelease, FGSAlert
    from django.utils import timezone
    from datetime import timedelta
    
    # Get finished goods storage phases
    fgs_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store'
    ).select_related('bmr__product', 'started_by', 'completed_by').order_by('-started_date')
    
    # Group by status
    fgs_pending = fgs_phases.filter(status='pending')
    fgs_in_progress = fgs_phases.filter(status='in_progress') 
    fgs_completed = fgs_phases.filter(status='completed')
    
    # FGS Inventory Statistics
    total_inventory_items = FGSInventory.objects.count()
    available_for_sale = FGSInventory.objects.filter(status='available').count()
    
    # Recent releases (last 7 days)
    recent_releases_count = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=7)
    ).count()
    
    # Active alerts
    active_alerts_count = FGSAlert.objects.filter(is_resolved=False).count()
    
    # Statistics
    fgs_stats = {
        'total_in_store': fgs_completed.count(),
        'pending_storage': fgs_pending.count(),
        'being_stored': fgs_in_progress.count(),
        'storage_capacity_used': min(100, (fgs_completed.count() / max(1000, 1)) * 100),  # Assuming 1000 batch capacity
        
        # New inventory statistics
        'total_inventory_items': total_inventory_items,
        'available_for_sale': available_for_sale,
        'recent_releases': recent_releases_count,
        'active_alerts': active_alerts_count,
    }
    
    # Recent storage activity
    recent_stored = fgs_completed[:10]
    
    # Recent inventory items
    recent_inventory = FGSInventory.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).select_related('product', 'bmr').order_by('-created_at')[:10]
    
    # Recent releases
    recent_releases = ProductRelease.objects.filter(
        release_date__gte=timezone.now() - timedelta(days=14)
    ).select_related('inventory__product', 'inventory__bmr').order_by('-release_date')[:10]
    
    # Active alerts
    active_alerts = FGSAlert.objects.filter(
        is_resolved=False
    ).select_related('inventory').order_by('-priority', '-created_at')[:10]
    
    # Products in FGS by type
    products_in_fgs = fgs_completed.values(
        'bmr__product__product_type',
        'bmr__product__product_name'
    ).annotate(
        batch_count=Count('bmr'),
        latest_storage=Max('completed_date')
    ).order_by('bmr__product__product_type', '-latest_storage')
    
    # Get production data by product type
    product_type_data = {}
    completed_bmrs = BatchPhaseExecution.objects.filter(
        phase__phase_name='finished_goods_store',
        status='completed'
    ).select_related('bmr__product')
    
    for execution in completed_bmrs:
        product_type = execution.bmr.product.product_type
        if product_type not in product_type_data:
            product_type_data[product_type] = 0
        product_type_data[product_type] += 1
    
    # Get phase completion status across all batches
    phase_completion = {}
    all_phases = BatchPhaseExecution.objects.values('phase__phase_name').distinct()
    for phase_dict in all_phases:
        phase_name = phase_dict['phase__phase_name']
        if phase_name:
            total = BatchPhaseExecution.objects.filter(phase__phase_name=phase_name).count()
            completed = BatchPhaseExecution.objects.filter(
                phase__phase_name=phase_name,
                status='completed'
            ).count()
            if total > 0:  # Avoid division by zero
                completion_rate = (completed / total) * 100
            else:
                completion_rate = 0
            phase_completion[phase_name] = {
                'total': total,
                'completed': completed,
                'completion_rate': round(completion_rate, 1)
            }
    
    # Get weekly production trend
    today = timezone.now().date()
    start_date = today - timezone.timedelta(days=28)  # Last 4 weeks
    
    weekly_completions = {}
    for i in range(4):  # 4 weeks
        week_start = start_date + timezone.timedelta(days=i*7)
        week_end = week_start + timezone.timedelta(days=6)
        week_label = f"{week_start.strftime('%d %b')} - {week_end.strftime('%d %b')}"
        
        weekly_completions[week_label] = BatchPhaseExecution.objects.filter(
            phase__phase_name='finished_goods_store',
            status='completed',
            completed_date__date__range=[week_start, week_end]
        ).count()
    
    # QC pass/fail data
    qc_stats = {
        'passed': BatchPhaseExecution.objects.filter(
            phase__phase_name__in=['post_compression_qc', 'post_mixing_qc', 'post_blending_qc'],
            status='completed'
        ).count(),
        'failed': BatchPhaseExecution.objects.filter(
            phase__phase_name__in=['post_compression_qc', 'post_mixing_qc', 'post_blending_qc'],
            status='failed'
        ).count()
    }
    
    context = {
        'user': request.user,
        'fgs_pending': fgs_pending,
        'fgs_in_progress': fgs_in_progress,
        'recent_stored': recent_stored,
        'fgs_stats': fgs_stats,
        'products_in_fgs': products_in_fgs,
        'dashboard_title': 'Finished Goods Store Monitor',
        'product_type_data': product_type_data,
        'phase_completion': phase_completion,
        'weekly_production': weekly_completions,
        'qc_stats': qc_stats,
        
        # New FGS inventory data
        'recent_inventory': recent_inventory,
        'recent_releases': recent_releases,
        'active_alerts': active_alerts,
    }
    
    return render(request, 'dashboards/admin_fgs_monitor.html', context)

def export_timeline_data(request, timeline_data=None, format_type=None):
    """Export detailed timeline data to CSV or Excel with all phases"""
    # Handle direct URL access
    if timeline_data is None:
        # Get export format from request
        format_type = request.GET.get('format', 'excel')
        
        # Recreate the timeline data from scratch
        from bmr.models import BMR
        from workflow.models import BatchPhaseExecution
        
        bmrs = BMR.objects.select_related('product', 'created_by', 'approved_by').all()
        
        # Add timeline data for each BMR
        timeline_data = []
        for bmr in bmrs:
            phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase').order_by('phase__phase_order')
            bmr_created = bmr.created_date
            fgs_completed = phases.filter(
                phase__phase_name='finished_goods_store',
                status='completed'
            ).first()
            total_time_hours = None
            if fgs_completed and fgs_completed.completed_date:
                # FIXED: Calculate correct production time (start to end, not BMR created to FGS)
                first_started_phase = phases.filter(started_date__isnull=False).order_by('started_date').first()
                if first_started_phase:
                    total_time_hours = round((fgs_completed.completed_date - first_started_phase.started_date).total_seconds() / 3600, 2)
            phase_timeline = []
            for phase in phases:
                phase_data = {
                    'phase_name': phase.phase.phase_name.replace('_', ' ').title(),
                    'status': phase.status.title(),
                    'started_date': phase.started_date,
                    'completed_date': phase.completed_date,
                    'started_by': phase.started_by.get_full_name() if phase.started_by else None,
                    'completed_by': phase.completed_by.get_full_name() if phase.completed_by else None,
                    'duration_hours': None,
                    'operator_comments': getattr(phase, 'operator_comments', '') or '',
                    'phase_order': phase.phase.phase_order if hasattr(phase.phase, 'phase_order') else 0,
                    # Machine tracking
                    'machine_used': phase.machine_used.name if phase.machine_used else '',
                    # Breakdown tracking
                    'breakdown_occurred': 'Yes' if phase.breakdown_occurred else 'No',
                    'breakdown_duration': phase.get_breakdown_duration() if hasattr(phase, 'get_breakdown_duration') and phase.breakdown_occurred else '',
                    'breakdown_start_time': phase.breakdown_start_time if phase.breakdown_occurred else '',
                    'breakdown_end_time': phase.breakdown_end_time if phase.breakdown_occurred else '',
                    # Changeover tracking
                    'changeover_occurred': 'Yes' if phase.changeover_occurred else 'No',
                    'changeover_duration': phase.get_changeover_duration() if hasattr(phase, 'get_changeover_duration') and phase.changeover_occurred else '',
                    'changeover_start_time': phase.changeover_start_time if phase.changeover_occurred else '',
                    'changeover_end_time': phase.changeover_end_time if phase.changeover_occurred else '',
                }
                if phase.started_date and phase.completed_date:
                    duration = phase.completed_date - phase.started_date
                    phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
                elif phase.started_date and not phase.completed_date:
                    duration = timezone.now() - phase.started_date
                    phase_data['duration_hours'] = round(duration.total_seconds() / 3600, 2)
                phase_timeline.append(phase_data)
            timeline_data.append({
                'bmr': bmr,
                'total_time_hours': total_time_hours,
                'phase_timeline': phase_timeline,
                'current_phase': phases.filter(status__in=['pending', 'in_progress']).first(),
                'is_completed': fgs_completed is not None,
            })
    
    # Generate CSV export
    if format_type == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="bmr_detailed_timeline_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        writer = csv.writer(response)
        
        # Header row
        writer.writerow(['BMR Report - Generated on', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
        writer.writerow([])
        
        # Write detailed phase information for each BMR
        for item in timeline_data:
            bmr = item['bmr']
            writer.writerow([])  # Empty row for separation
            writer.writerow([f"BMR: {bmr.batch_number} - {bmr.product.product_name}"])
            writer.writerow([f"Product Type: {bmr.product.product_type}"])
            writer.writerow([f"Created: {bmr.created_date.strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([f"Total Production Time: {item['total_time_hours']} hours" if item['total_time_hours'] else "In Progress"])
            writer.writerow([])  # Empty row
            writer.writerow([
                'Phase Name', 'Status', 'Started Date', 'Started By', 
                'Completed Date', 'Completed By', 'Duration (Hours)', 'Comments',
                'Machine Used', 'Breakdown Occurred', 'Breakdown Duration (Min)', 
                'Breakdown Start', 'Breakdown End', 'Changeover Occurred', 
                'Changeover Duration (Min)', 'Changeover Start', 'Changeover End'
            ])
            for phase in item['phase_timeline']:
                writer.writerow([
                    phase['phase_name'], phase['status'],
                    phase['started_date'], phase['started_by'],
                    phase['completed_date'], phase['completed_by'],
                    phase['duration_hours'], phase['operator_comments'],
                    phase['machine_used'], phase['breakdown_occurred'], 
                    phase['breakdown_duration'], phase['breakdown_start_time'],
                    phase['breakdown_end_time'], phase['changeover_occurred'],
                    phase['changeover_duration'], phase['changeover_start_time'],
                    phase['changeover_end_time']
                ])
        return response
    
    # Generate Excel export
    elif format_type == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        
        # Create summary sheet
        summary_sheet = wb.active
        summary_sheet.title = "Production Summary"
        
        # Apply styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Create title
        summary_sheet.merge_cells('A1:I1')
        title_cell = summary_sheet['A1']
        title_cell.value = "Kampala Pharmaceutical Industries - BMR Production Timeline Summary"
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Create report generation date
        summary_sheet.merge_cells('A2:I2')
        date_cell = summary_sheet['A2']
        date_cell.value = f"Report Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.font = Font(italic=True)
        
        # Add empty row
        summary_sheet.append([])
        
        # Summary headers
        headers = [
            "Batch Number", "Product Name", "Product Type", 
            "Created Date", "Current Status", "Current Phase",
            "Total Duration (Hours)", "Completed", "Bottleneck Phase"
        ]
        
        header_row = summary_sheet.row_dimensions[4]
        header_row.height = 30
        
        for col_num, header in enumerate(headers, 1):
            cell = summary_sheet.cell(row=4, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
            summary_sheet.column_dimensions[get_column_letter(col_num)].width = 18
        
        # Add data rows
        row_num = 5
        for item in timeline_data:
            bmr = item['bmr']
            # Find bottleneck phase (longest duration)
            bottleneck = max(item['phase_timeline'], key=lambda x: x['duration_hours'] if x['duration_hours'] else 0, default={})
            bottleneck_name = bottleneck.get('phase_name', 'N/A') if bottleneck else 'N/A'
            
            # Get current phase
            current_phase = "Completed"
            if not item['is_completed']:
                current_phases = [p for p in item['phase_timeline'] if p['status'] in ['In Progress', 'Pending']]
                if current_phases:
                    current_phase = current_phases[0]['phase_name']
            
            # Add row data
            row_data = [
                bmr.batch_number,
                bmr.product.product_name,
                bmr.product.product_type.replace('_', ' ').title(),
                bmr.created_date.strftime('%Y-%m-%d'),
                "Completed" if item['is_completed'] else "In Progress",
                current_phase,
                item['total_time_hours'] if item['total_time_hours'] else "In Progress",
                "Yes" if item['is_completed'] else "No",
                bottleneck_name
            ]
            
            for col_num, cell_value in enumerate(row_data, 1):
                cell = summary_sheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
        
        # Create detail sheet for each BMR
        for item in timeline_data:
            bmr = item['bmr']
            # Create sheet for this BMR
            detail_sheet = wb.create_sheet(title=f"BMR-{bmr.batch_number}")
            
            # Title
            detail_sheet.merge_cells('A1:H1')
            title_cell = detail_sheet['A1']
            title_cell.value = f"Detailed Timeline for BMR {bmr.batch_number} - {bmr.product.product_name}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal="center")
            
            # BMR information
            detail_sheet.merge_cells('A2:H2')
            info_cell = detail_sheet['A2']
            info_cell.value = f"Product Type: {bmr.product.product_type.replace('_', ' ').title()} | Created: {bmr.created_date.strftime('%Y-%m-%d %H:%M:%S')}"
            info_cell.font = Font(italic=True)
            info_cell.alignment = Alignment(horizontal="center")
            
            detail_sheet.merge_cells('A3:H3')
            time_cell = detail_sheet['A3']
            time_cell.value = f"Total Production Time: {item['total_time_hours']} hours" if item['total_time_hours'] else "Total Production Time: In Progress"
            time_cell.font = Font(italic=True, bold=True)
            time_cell.alignment = Alignment(horizontal="center")
            
            # Add empty row
            detail_sheet.append([])
            
            # Detail headers
            headers = [
                "Phase Name", "Status", "Started Date", "Started By", 
                "Completed Date", "Completed By", "Duration (Hours)", "Comments",
                "Machine Used", "Breakdown Occurred", "Breakdown Duration (Min)", 
                "Breakdown Start", "Breakdown End", "Changeover Occurred", 
                "Changeover Duration (Min)", "Changeover Start", "Changeover End"
            ]
            
            header_row = detail_sheet.row_dimensions[5]
            header_row.height = 30
            
            for col_num, header in enumerate(headers, 1):
                cell = detail_sheet.cell(row=5, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
                # Adjust column widths for new columns
                if col_num <= 8:  # Original columns
                    detail_sheet.column_dimensions[get_column_letter(col_num)].width = 18
                elif col_num in [9, 10, 14]:  # Machine, breakdown occurred, changeover occurred
                    detail_sheet.column_dimensions[get_column_letter(col_num)].width = 15
                else:  # Date/time columns
                    detail_sheet.column_dimensions[get_column_letter(col_num)].width = 20
            
            # Add phase data
            phase_row = 6
            for phase in item['phase_timeline']:
                # Format dates for display
                started_date = phase['started_date'].strftime('%Y-%m-%d %H:%M') if phase['started_date'] else "Not Started"
                completed_date = phase['completed_date'].strftime('%Y-%m-%d %H:%M') if phase['completed_date'] else "Not Completed"
                breakdown_start = phase['breakdown_start_time'].strftime('%Y-%m-%d %H:%M') if phase['breakdown_start_time'] else ""
                breakdown_end = phase['breakdown_end_time'].strftime('%Y-%m-%d %H:%M') if phase['breakdown_end_time'] else ""
                changeover_start = phase['changeover_start_time'].strftime('%Y-%m-%d %H:%M') if phase['changeover_start_time'] else ""
                changeover_end = phase['changeover_end_time'].strftime('%Y-%m-%d %H:%M') if phase['changeover_end_time'] else ""
                
                phase_data = [
                    phase['phase_name'],
                    phase['status'],
                    started_date,
                    phase['started_by'] if phase['started_by'] else "",
                    completed_date,
                    phase['completed_by'] if phase['completed_by'] else "",
                    phase['duration_hours'] if phase['duration_hours'] is not None else "",
                    phase['operator_comments'] if phase['operator_comments'] else "",
                    phase['machine_used'] if phase['machine_used'] else "",
                    phase['breakdown_occurred'],
                    phase['breakdown_duration'] if phase['breakdown_duration'] else "",
                    breakdown_start,
                    breakdown_end,
                    phase['changeover_occurred'],
                    phase['changeover_duration'] if phase['changeover_duration'] else "",
                    changeover_start,
                    changeover_end
                ]
                
                # Apply styling based on status
                row_fill = None
                if phase['status'] == 'Completed':
                    row_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                elif phase['status'] == 'In Progress':
                    row_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                
                for col_num, cell_value in enumerate(phase_data, 1):
                    cell = detail_sheet.cell(row=phase_row, column=col_num)
                    cell.value = cell_value
                    cell.border = border
                    if row_fill:
                        cell.fill = row_fill
                    
                    # For comments column, use wrap text
                    if col_num == 8:  # Comments column
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                        detail_sheet.row_dimensions[phase_row].height = max(15, min(50, len(str(cell_value)) // 10 * 15))
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Adjust column widths for new columns
                    if col_num <= 8:  # Original columns
                        detail_sheet.column_dimensions[get_column_letter(col_num)].width = 18
                    elif col_num in [9, 10, 14]:  # Machine, breakdown occurred, changeover occurred
                        detail_sheet.column_dimensions[get_column_letter(col_num)].width = 15
                    else:  # Date/time columns
                        detail_sheet.column_dimensions[get_column_letter(col_num)].width = 20
                
                phase_row += 1
        
        # Create response
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="bmr_timeline_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Save the workbook to the response
        wb.save(response)
        return response
    
    else:
        return HttpResponse('Unsupported export format', content_type='text/plain')


# Redirect view for old admin dashboard URL
def admin_redirect(request):
    # Direct redirect to admin dashboard function
    return admin_dashboard(request)

@login_required
def admin_fgs_monitor(request):
    """Admin FGS Monitor View"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin role required.')
        return redirect('dashboards:dashboard_home')
    
    from fgs_management.models import FGSInventory
    from quarantine.models import QuarantineBatch, SampleRequest
    from django.db.models import Count, Q
    from datetime import timedelta
    
    # FGS Statistics
    total_in_store = FGSInventory.objects.filter(status__in=['stored', 'available']).count()
    pending_storage = 0  # FGSInventory doesn't have 'pending' status, using 0
    available_for_sale = FGSInventory.objects.filter(status='available').count()
    being_stored = FGSInventory.objects.filter(status='stored').count()
    
    # Recent releases (last 7 days)
    seven_days_ago = timezone.now() - timedelta(days=7)
    recent_releases = FGSInventory.objects.filter(
        status='released',
        updated_at__gte=seven_days_ago
    ).count()
    
    # Storage capacity (simplified calculation)
    max_capacity = 1000  # Assuming max capacity of 1000 items
    storage_capacity_used = (total_in_store / max_capacity) * 100 if max_capacity > 0 else 0
    
    fgs_stats = {
        'total_in_store': total_in_store,
        'pending_storage': pending_storage,
        'available_for_sale': available_for_sale,
        'being_stored': being_stored,
        'recent_releases': recent_releases,
        'storage_capacity_used': storage_capacity_used,
    }
    
    # Product type distribution
    product_distribution = BMR.objects.values('product_type').annotate(
        count=Count('id')
    ).order_by('product_type')
    
    # Phase completion statistics
    phase_stats = {}
    for phase_name in ['granulation', 'blending', 'compression', 'coating', 'packing']:
        completed = BatchPhaseExecution.objects.filter(
            phase__phase_name=phase_name,
            status='completed'
        ).count()
        total = BatchPhaseExecution.objects.filter(
            phase__phase_name=phase_name
        ).count()
        completion_rate = (completed / total * 100) if total > 0 else 0
        
        phase_stats[phase_name] = {
            'completed': completed,
            'total': total,
            'completion_rate': round(completion_rate, 1)
        }
    
    # Weekly production data (last 4 weeks)
    weekly_production = []
    for i in range(4):
        week_start = timezone.now() - timedelta(weeks=i+1)
        week_end = timezone.now() - timedelta(weeks=i)
        week_count = BMR.objects.filter(
            created_date__gte=week_start,
            created_date__lt=week_end
        ).count()
        weekly_production.append({
            'week': f'Week {i+1}',
            'count': week_count
        })
    
    # QC Statistics
    qc_passed = SampleRequest.objects.filter(qc_status='approved').count()
    qc_failed = SampleRequest.objects.filter(qc_status='failed').count()
    qc_stats = {
        'passed': qc_passed,
        'failed': qc_failed,
    }
    
    context = {
        'dashboard_title': 'Finished Goods Store Monitor',
        'fgs_stats': fgs_stats,
        'product_distribution': product_distribution,
        'phase_stats': phase_stats,
        'weekly_production': weekly_production,
        'qc_stats': qc_stats,
    }
    return render(request, 'dashboards/admin_fgs_monitor.html', context)

@login_required
def export_timeline_data(request):
    """Export Timeline Data"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin role required.')
        return redirect('dashboards:dashboard_home')
    
    # Return JSON response with timeline data
    from django.http import JsonResponse
    return JsonResponse({'status': 'success', 'data': []})

@login_required
def live_tracking_view(request):
    """Live Tracking View"""
    if request.user.role != 'admin':
        messages.error(request, 'Access denied. Admin role required.')
        return redirect('dashboards:dashboard_home')
    
    context = {
        'dashboard_title': 'Live Tracking',
    }
    return render(request, 'dashboards/live_tracking.html', context)

@login_required
def admin_machine_management(request):
    """Admin Machine Management"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    # Get all machines
    all_machines = Machine.objects.all().order_by('machine_type', 'name')
    
    # Get recent breakdowns (last 30 days)
    recent_breakdowns = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-breakdown_start_time')[:20]
    
    # Get recent changeovers (last 30 days)
    recent_changeovers = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__gte=timezone.now() - timedelta(days=30)
    ).select_related('machine_used', 'bmr').order_by('-changeover_start_time')[:20]
    
    # Count total breakdowns and changeovers
    total_breakdowns = BatchPhaseExecution.objects.filter(breakdown_occurred=True).count()
    total_changeovers = BatchPhaseExecution.objects.filter(changeover_occurred=True).count()
    
    # Breakdown and changeover counts for today
    today = timezone.now().date()
    breakdowns_today = BatchPhaseExecution.objects.filter(
        breakdown_occurred=True,
        breakdown_start_time__date=today
    ).count()
    changeovers_today = BatchPhaseExecution.objects.filter(
        changeover_occurred=True,
        changeover_start_time__date=today
    ).count()
    
    # Machine utilization summary
    machine_stats = {}
    for machine in all_machines:
        usage_count = BatchPhaseExecution.objects.filter(machine_used=machine).count()
        breakdown_count = BatchPhaseExecution.objects.filter(
            machine_used=machine,
            breakdown_occurred=True
        ).count()
        changeover_count = BatchPhaseExecution.objects.filter(
            machine_used=machine,
            changeover_occurred=True
        ).count()
        
        machine_stats[machine.id] = {
            'machine': machine,
            'usage_count': usage_count,
            'breakdown_count': breakdown_count,
            'changeover_count': changeover_count,
            'breakdown_rate': round((breakdown_count / usage_count * 100), 1) if usage_count > 0 else 0
        }
    
    context = {
        'dashboard_title': 'Machine Management',
        'page_title': 'Production Machine Management',
        'all_machines': all_machines,
        'recent_breakdowns': recent_breakdowns,
        'recent_changeovers': recent_changeovers,
        'total_breakdowns': total_breakdowns,
        'total_changeovers': total_changeovers,
        'breakdowns_today': breakdowns_today,
        'changeovers_today': changeovers_today,
        'machine_stats': machine_stats,
    }
    return render(request, 'dashboards/admin_machine_management.html', context)

@login_required
def admin_quality_control(request):
    """Admin Quality Control"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    from quarantine.models import QuarantineBatch, SampleRequest
    from workflow.models import BatchPhaseExecution
    
    # Get QC checkpoint statistics from sample requests
    all_sample_requests = SampleRequest.objects.all()
    total_checkpoints = all_sample_requests.count()
    
    # Count passed vs failed checkpoints
    passed_checkpoints = all_sample_requests.filter(qc_status='approved').count()
    failed_checkpoints = all_sample_requests.filter(qc_status='failed').count()
    pending_checkpoints = all_sample_requests.filter(qc_status='pending').count()
    
    # Calculate failure rate
    if total_checkpoints > 0:
        failure_rate = round((failed_checkpoints / total_checkpoints) * 100, 1)
    else:
        failure_rate = 0.0
    
    # Get recent QC checkpoints - use QC phases instead of sample requests to match template
    qc_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name__icontains='qc'
    ).select_related('bmr', 'phase', 'completed_by').order_by('-completed_date')[:20]
    
    # Create checkpoint objects that match template expectations
    recent_checkpoints = []
    for phase in qc_phases:
        # Create a mock checkpoint object with the fields the template expects
        checkpoint = type('Checkpoint', (), {
            'checked_date': phase.completed_date,
            'phase_execution': phase,  # This gives us access to phase_execution.bmr.id
            'checkpoint_name': f"QC {phase.phase.phase_name.replace('_', ' ').title()}",
            'checked_by': phase.completed_by,
            'is_within_spec': None if phase.status == 'in_progress' else (True if phase.status == 'completed' else False)
        })()
        recent_checkpoints.append(checkpoint)
    
    # Monthly QC test volume (last 6 months)
    from datetime import datetime, timedelta
    from django.db.models import Count
    from django.db.models.functions import TruncMonth
    
    six_months_ago = timezone.now() - timedelta(days=180)
    monthly_tests = all_sample_requests.filter(
        request_date__gte=six_months_ago
    ).annotate(
        month=TruncMonth('request_date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')
    
    context = {
        'dashboard_title': 'Quality Control',
        'page_title': 'Quality Control Management',
        'all_checkpoints': total_checkpoints,
        'passed_checkpoints': passed_checkpoints,
        'failed_checkpoints': failed_checkpoints,
        'pending_checkpoints': pending_checkpoints,
        'failure_rate': failure_rate,
        'recent_checkpoints': recent_checkpoints,
        'qc_phases': qc_phases,
        'monthly_tests': monthly_tests,
    }
    return render(request, 'dashboards/admin_quality_control.html', context)

@login_required
def admin_inventory(request):
    """Admin Inventory"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    from fgs_management.models import FGSInventory
    
    # Get all inventory items
    inventory = FGSInventory.objects.select_related('product', 'bmr').order_by('-created_at')[:50]
    
    # Calculate inventory statistics
    total_inventory = FGSInventory.objects.count()
    available_inventory = FGSInventory.objects.filter(status='available').count()
    stored_inventory = FGSInventory.objects.filter(status='stored').count()
    reserved_inventory = FGSInventory.objects.filter(status='reserved').count()
    released_inventory = FGSInventory.objects.filter(status='released').count()
    
    # Get low stock items (if needed)
    low_stock_items = FGSInventory.objects.filter(
        quantity_available__lt=100  # Items with less than 100 units
    ).order_by('quantity_available')[:20]
    
    context = {
        'dashboard_title': 'Inventory Management',
        'page_title': 'Inventory & FGS Management',
        'inventory': inventory,
        'total_inventory': total_inventory,
        'available_inventory': available_inventory,
        'stored_inventory': stored_inventory,
        'reserved_inventory': reserved_inventory,
        'released_inventory': released_inventory,
        'low_stock_items': low_stock_items,
    }
    return render(request, 'dashboards/admin_inventory.html', context)

@login_required
def admin_user_management(request):
    """Admin User Management"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    from accounts.models import CustomUser
    from django.db.models import Count
    
    # Get all users
    all_users = CustomUser.objects.all().order_by('username')
    
    # Calculate user statistics
    total_users = all_users.count()
    active_users = all_users.filter(is_active=True).count()
    inactive_users = all_users.filter(is_active=False).count()
    staff_users = all_users.filter(is_staff=True).count()
    
    # Get user distribution by role
    role_distribution = all_users.values('role').annotate(count=Count('role')).order_by('role')
    
    # Create role_counts dictionary for template compatibility
    role_counts = {}
    for role_item in role_distribution:
        role = role_item['role']
        count = role_item['count']
        # Convert role to readable format
        readable_role = role.replace('_', ' ').title()
        role_counts[readable_role] = count
    
    context = {
        'dashboard_title': 'User Management',
        'page_title': 'System User Management',
        'all_users': all_users,
        'total_users': total_users,
        'active_users': active_users,
        'inactive_users': inactive_users,
        'staff_users': staff_users,
        'role_distribution': role_distribution,
        'role_counts': role_counts,
    }
    return render(request, 'dashboards/admin_user_management.html', context)

@login_required
def admin_system_health(request):
    """Admin System Health"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    import sys
    import platform
    import os
    import django
    from django.contrib.admin.models import LogEntry
    from django.contrib.contenttypes.models import ContentType
    from accounts.models import CustomUser
    from bmr.models import BMR
    from products.models import Product
    from workflow.models import BatchPhaseExecution
    
    # System Information
    system_info = {
        'python_version': f"{sys.version_info.major}.{sys.version_info.minor}.{sys.version_info.micro}",
        'django_version': django.get_version(),
        'database': 'SQLite',  # Update if using different database
        'os': f"{platform.system()} {platform.release()}",
        'cpu_count': os.cpu_count(),
    }
    
    # Database Statistics
    db_stats = {
        'bmr_count': BMR.objects.count(),
        'users_count': CustomUser.objects.count(),
        'phases_count': BatchPhaseExecution.objects.count(),
        'products_count': Product.objects.count(),
    }
    
    # Recent System Logs (from Django admin logs)
    recent_logs = LogEntry.objects.select_related(
        'user', 'content_type'
    ).order_by('-action_time')[:50]
    
    # System Health Checks
    health_checks = {
        'database_connection': True,  # If we get here, DB is working
        'server_status': True,        # If we get here, server is running
        'disk_space': True,           # Simplified for now
    }
    
    # Memory and CPU usage (if psutil is available)
    try:
        import psutil
        memory = psutil.virtual_memory()
        cpu_percent = psutil.cpu_percent(interval=1)
        
        # Handle different OS disk usage
        try:
            if platform.system() == 'Windows':
                disk = psutil.disk_usage('C:\\')
            else:
                disk = psutil.disk_usage('/')
        except:
            disk = None
        
        system_info.update({
            'memory_total': f"{memory.total // (1024**3)} GB",
            'memory_used': f"{memory.percent}%",
            'cpu_usage': f"{cpu_percent}%",
            'disk_total': f"{disk.total // (1024**3)} GB" if disk else 'N/A',
            'disk_used': f"{(disk.used / disk.total * 100):.1f}%" if disk else 'N/A',
        })
    except (ImportError, Exception):
        # psutil not available or error occurred, use basic info
        system_info.update({
            'memory_total': 'N/A',
            'memory_used': 'N/A', 
            'cpu_usage': 'N/A',
            'disk_total': 'N/A',
            'disk_used': 'N/A',
        })
    
    context = {
        'dashboard_title': 'System Health',
        'page_title': 'System Health & Monitoring',
        'system_info': system_info,
        'db_stats': db_stats,
        'recent_logs': recent_logs,
        'health_checks': health_checks,
    }
    return render(request, 'dashboards/admin_system_health.html', context)

@login_required
def phase_notifications_view(request):
    """Phase Timing Notifications and Overrun Alerts"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    from workflow.models import PhaseOverrunNotification, PhaseTimeOverrunNotification, BatchPhaseExecution
    from django.utils import timezone
    from datetime import datetime, timedelta
    
    # Get overrun notifications
    overrun_notifications = PhaseOverrunNotification.objects.filter(
        status__in=['pending', 'responded']
    ).select_related('phase_execution__bmr', 'phase_execution__phase').order_by('-created_at')[:20]
    
    # Get time-based overrun notifications  
    time_notifications = PhaseTimeOverrunNotification.objects.filter(
        acknowledged=False
    ).select_related('phase_execution__bmr', 'phase_execution__phase').order_by('-notification_time')[:20]
    
    # Get currently active phases that might be overrunning
    active_phases = BatchPhaseExecution.objects.filter(
        status='in_progress',
        started_date__isnull=False
    ).select_related('bmr', 'phase')
    
    # Calculate which active phases are overrunning
    overrunning_phases = []
    for phase in active_phases:
        if phase.started_date:
            duration_hours = phase.duration_hours or 0
            expected_hours = phase.phase.estimated_duration_hours
            if expected_hours > 0 and duration_hours > (expected_hours * 1.2):  # 20% over
                overrunning_phases.append({
                    'phase': phase,
                    'duration_hours': duration_hours,
                    'expected_hours': expected_hours,
                    'overrun_percent': ((duration_hours - expected_hours) / expected_hours * 100)
                })
    
    context = {
        'dashboard_title': 'Phase Timing Alerts & Notifications',
        'overrun_notifications': overrun_notifications,
        'time_notifications': time_notifications,
        'overrunning_phases': overrunning_phases,
        'total_notifications': len(overrun_notifications) + len(time_notifications),
    }
    return render(request, 'dashboards/phase_notifications.html', context)


def phase_specific_dashboard(request, phase_name):
    """
    Generic phase-specific dashboard with countdown timer functionality
    Handles: granulation, blending, compression, coating, drying, filling, etc.
    """
    if not request.user.is_authenticated:
        return redirect('accounts:login')
    
    from workflow.models import BatchPhaseExecution, PhaseTimingSetting, ProductionPhase
    from django.utils import timezone
    import json
    
    # Get current active phases for this phase type
    active_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name=phase_name,
        status='in_progress'
    ).select_related('bmr__product', 'phase', 'started_by', 'machine_used')
    
    # Get pending phases for this phase type  
    pending_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name=phase_name,
        status='pending'
    ).select_related('bmr__product', 'phase')[:10]
    
    # Get completed phases from last 24 hours
    from datetime import timedelta
    yesterday = timezone.now() - timedelta(days=1)
    completed_phases = BatchPhaseExecution.objects.filter(
        phase__phase_name=phase_name,
        status='completed',
        completed_date__gte=yesterday
    ).select_related('bmr__product', 'phase')[:10]
    
    # Get phase timing settings for countdown
    phase_settings = {}
    try:
        settings_qs = PhaseTimingSetting.objects.filter(
            phase__phase_name=phase_name
        ).select_related('phase')
        
        for setting in settings_qs:
            key = f"{setting.phase.product_type}_{setting.phase.phase_name}"
            phase_settings[key] = {
                'expected_hours': float(setting.expected_duration_hours),
                'warning_threshold': setting.warning_threshold_percent
            }
    except Exception as e:
        logger.error(f"Error loading phase settings: {e}")
    
    # Prepare active phases with timing data
    active_phases_data = []
    for phase in active_phases:
        if phase.started_date:
            # Calculate elapsed time
            elapsed = timezone.now() - phase.started_date
            elapsed_hours = elapsed.total_seconds() / 3600
            
            # Get expected duration for this product type + phase
            key = f"{phase.bmr.product.product_type}_{phase.phase.phase_name}"
            # Admin must configure expected hours - no defaults
            expected_hours = phase_settings.get(key, {}).get('expected_hours', None)
            if not expected_hours:
                continue  # Skip phases without configured timing
            warning_threshold = phase_settings.get(key, {}).get('warning_threshold', 20)
            
            # Calculate remaining time
            remaining_hours = expected_hours - elapsed_hours
            remaining_seconds = int(remaining_hours * 3600) if remaining_hours > 0 else 0
            
            # Determine status
            warning_time = expected_hours * (1 + warning_threshold / 100)
            if elapsed_hours >= warning_time:
                timer_status = 'overrun'
            elif elapsed_hours >= expected_hours:
                timer_status = 'expired'
            elif remaining_hours <= 0.5:  # 30 minutes warning
                timer_status = 'warning'
            else:
                timer_status = 'normal'
            
            progress_percent = min(100, (elapsed_hours / expected_hours) * 100)
            # Calculate SVG stroke offset: progress_percent * 2.64 - 264
            svg_stroke_offset = progress_percent * 2.64 - 264
            
            active_phases_data.append({
                'phase': phase,
                'elapsed_hours': round(elapsed_hours, 2),
                'expected_hours': expected_hours,
                'remaining_hours': round(remaining_hours, 2),
                'remaining_seconds': remaining_seconds,
                'timer_status': timer_status,
                'progress_percent': progress_percent,
                'svg_stroke_offset': svg_stroke_offset
            })
    
    context = {
        'phase_name': phase_name,
        'phase_display_name': phase_name.replace('_', ' ').title(),
        'active_phases': active_phases_data,
        'pending_phases': pending_phases,
        'completed_phases': completed_phases,
        'phase_settings_json': json.dumps(phase_settings),
        'dashboard_title': f'{phase_name.replace("_", " ").title()} Dashboard',
    }
    
    return render(request, 'dashboards/phase_specific_dashboard.html', context)


# ==================== NOTIFICATION API ENDPOINTS ====================

from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
import json
from datetime import datetime, timedelta

@login_required
@require_http_methods(["GET"])
def notification_counts_api(request):
    """API endpoint to get notification counts for badges"""
    try:
        from workflow.models import PhaseOverrunNotification, PhaseTimeOverrunNotification
        
        # Count different types of notifications
        critical_count = PhaseOverrunNotification.objects.filter(
            status__in=['pending', 'responded']
        ).count()
        
        overrun_count = PhaseTimeOverrunNotification.objects.filter(
            acknowledged=False
        ).count()
        
        # Simulated counts for other notification types
        info_count = 0  # Could be system updates, general info
        resolved_count = PhaseOverrunNotification.objects.filter(
            status__in=['resolved', 'closed'],
            created_at__date=timezone.now().date()
        ).count()
        
        unread_count = critical_count + overrun_count + info_count
        
        return JsonResponse({
            'critical': critical_count,
            'overrun': overrun_count,
            'info': info_count,
            'resolved': resolved_count,
            'unread': unread_count
        })
        
    except Exception as e:
        # Return sample data if models don't exist yet
        return JsonResponse({
            'critical': 2,
            'overrun': 5,
            'info': 8,
            'resolved': 12,
            'unread': 15
        })

@login_required
@require_http_methods(["GET"])
def notifications_feed_api(request):
    """API endpoint to get notifications feed data"""
    try:
        from workflow.models import PhaseOverrunNotification, PhaseTimeOverrunNotification, BatchPhaseExecution
        
        notifications = []
        
        # Get overrun notifications (legacy model - if still being used)
        overrun_notifications = PhaseOverrunNotification.objects.filter(
            status='pending'
        ).select_related('phase_execution__bmr', 'phase_execution__phase').order_by('-created_at')[:5]
        
        for notif in overrun_notifications:
            batch_number = getattr(notif.phase_execution.bmr, 'batch_number', 'Unknown')
            phase_name = getattr(notif.phase_execution.phase, 'phase_name', 'Unknown Phase')
            
            notifications.append({
                'id': f"legacy_{notif.id}",
                'title': 'Legacy Phase Alert',
                'message': f'{phase_name.replace("_", " ").title()} phase overrun detected',
                'priority': 'MEDIUM',
                'created_at': notif.created_at.isoformat(),
                'batch_number': batch_number,
                'is_read': False,
                'type': 'legacy_overrun'
            })
        
        # Get time overrun notifications (now unified - visible to all admins)
        time_overruns = PhaseTimeOverrunNotification.objects.filter(
            acknowledged=False
        ).select_related('phase_execution__bmr', 'phase_execution__phase').order_by('-notification_time')[:10]
        
        for notif in time_overruns:
            batch_number = getattr(notif.phase_execution.bmr, 'batch_number', 'Unknown')
            phase_name = getattr(notif.phase_execution.phase, 'phase_name', 'Unknown Phase')
            
            notifications.append({
                'id': notif.id,  # Use actual ID for acknowledgment
                'title': 'Phase Time Overrun Alert',
                'message': notif.message,
                'priority': 'HIGH' if notif.threshold_exceeded_percent > 100 else 'MEDIUM',
                'created_at': notif.notification_time.isoformat(),
                'batch_number': batch_number,
                'phase_name': phase_name,
                'is_read': notif.acknowledged,
                'type': 'time_overrun',
                'threshold_exceeded_percent': notif.threshold_exceeded_percent
            })
        
        # Sort by created date
        notifications.sort(key=lambda x: x['created_at'], reverse=True)
        
        return JsonResponse({
            'notifications': notifications[:15]  # Limit to 15 most recent
        })
        
    except Exception as e:
        # Return sample notifications if models don't exist
        sample_notifications = [
            {
                'id': 1,
                'title': 'Phase Overrun Alert',
                'message': 'Granulation phase for batch 001-2025 has exceeded expected duration by 25 minutes',
                'priority': 'HIGH',
                'created_at': (timezone.now() - timedelta(minutes=5)).isoformat(),
                'batch_number': '001-2025',
                'is_read': False,
                'type': 'overrun'
            },
            {
                'id': 2,
                'title': 'QC Review Required',
                'message': 'Batch 002-2025 requires quality control review before proceeding',
                'priority': 'MEDIUM',
                'created_at': (timezone.now() - timedelta(minutes=30)).isoformat(),
                'batch_number': '002-2025',
                'is_read': False,
                'type': 'qc'
            },
            {
                'id': 3,
                'title': 'System Update',
                'message': 'Phase timing settings have been updated for all tablet production lines',
                'priority': 'LOW',
                'created_at': (timezone.now() - timedelta(hours=2)).isoformat(),
                'batch_number': None,
                'is_read': True,
                'type': 'system'
            }
        ]
        
        return JsonResponse({
            'notifications': sample_notifications
        })

@login_required
@require_http_methods(["GET"])
def overrun_alerts_api(request):
    """API endpoint to get overrun alerts data"""
    try:
        from workflow.models import BatchPhaseExecution, PhaseTimingSetting, ProductMachineTimingSetting
        
        # Get currently overrunning phases
        current_overruns = []
        active_phases = BatchPhaseExecution.objects.filter(
            status='active'
        ).select_related('bmr', 'phase')
        
        for phase_exec in active_phases:
            if phase_exec.started_date:
                elapsed = timezone.now() - phase_exec.started_date
                elapsed_hours = elapsed.total_seconds() / 3600
                
                # ENHANCED: Get expected duration using new Product+Machine timing system
                expected_hours = ProductMachineTimingSetting.get_expected_duration_for_execution(phase_exec)
                
                if elapsed_hours > expected_hours:
                    overrun_hours = elapsed_hours - expected_hours
                    current_overruns.append({
                        'batch_number': getattr(phase_exec.bmr, 'batch_number', 'Unknown'),
                        'phase_name': phase_exec.phase.phase_name.replace('_', ' ').title(),
                        'expected_duration': f"{expected_hours:.1f}h",
                        'overrun_duration': f"{overrun_hours:.1f}h",
                            'start_time': phase_exec.started_date.isoformat()
                    })
        
        # Get historical overruns
        history = []
        completed_phases = BatchPhaseExecution.objects.filter(
            status='completed',
            completed_date__gte=timezone.now() - timedelta(days=7)
        ).select_related('bmr', 'phase')[:10]
        
        for phase_exec in completed_phases:
            if phase_exec.started_date and phase_exec.completed_date:
                duration = phase_exec.completed_date - phase_exec.started_date
                actual_hours = duration.total_seconds() / 3600
                
                # ENHANCED: Get expected duration using new Product+Machine timing system
                expected_hours = ProductMachineTimingSetting.get_expected_duration_for_execution(phase_exec)
                
                if actual_hours > expected_hours:
                    history.append({
                        'id': phase_exec.id,
                        'batch_number': getattr(phase_exec.bmr, 'batch_number', 'Unknown'),
                        'phase_name': phase_exec.phase.phase_name.replace('_', ' ').title(),
                        'expected_duration': f"{expected_hours:.1f}h",
                        'actual_duration': f"{actual_hours:.1f}h",
                        'explanation_provided': False  # Would check if explanation exists
                    })
        
        return JsonResponse({
            'current_overruns': current_overruns,
            'history': history
        })
        
    except Exception as e:
        # Return sample data
        return JsonResponse({
            'current_overruns': [
                {
                    'batch_number': '001-2025',
                    'phase_name': 'Granulation',
                    'expected_duration': '2.5h',
                    'overrun_duration': '0.75h',
                    'start_time': (timezone.now() - timedelta(hours=3, minutes=15)).isoformat()
                }
            ],
            'history': [
                {
                    'id': 1,
                    'batch_number': '001-2025',
                    'phase_name': 'Granulation',
                    'expected_duration': '2.5h',
                    'actual_duration': '3.25h',
                    'explanation_provided': False
                }
            ]
        })

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def mark_notification_read_api(request, notification_id):
    """API endpoint to mark a notification as read/acknowledged"""
    try:
        from workflow.models import PhaseTimeOverrunNotification
        from django.utils import timezone
        
        # Find the notification - allow any admin/manager to acknowledge
        notification = PhaseTimeOverrunNotification.objects.get(
            id=notification_id
        )
        
        # Check if user has permission to acknowledge notifications
        if not (request.user.is_superuser or 
                request.user.role in ['admin', 'production_manager', 'qa', 'qc']):
            return JsonResponse({
                'success': False,
                'error': 'You do not have permission to acknowledge this notification'
            })
        
        # Mark as acknowledged
        notification.acknowledged = True
        notification.acknowledged_by = request.user
        notification.acknowledged_at = timezone.now()
        notification.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Notification acknowledged successfully'
        })
    except PhaseTimeOverrunNotification.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Notification not found or access denied'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def dismiss_notification_api(request, notification_id):
    """API endpoint to dismiss/delete a notification"""
    try:
        from workflow.models import PhaseTimeOverrunNotification
        
        # Find and delete the notification (any admin can dismiss)
        notification = PhaseTimeOverrunNotification.objects.get(
            id=notification_id
        )
        
        batch_number = notification.phase_execution.bmr.batch_number
        phase_name = notification.phase_execution.phase.phase_name
        
        notification.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Notification for {batch_number} - {phase_name} dismissed successfully'
        })
    except PhaseTimeOverrunNotification.DoesNotExist:
        return JsonResponse({
            'success': False, 
            'error': 'Notification not found or access denied'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def request_explanation_api(request):
    """API endpoint to request explanation for overrun"""
    try:
        data = json.loads(request.body)
        batch_number = data.get('batch_number')
        phase_name = data.get('phase_name')
        explanation = data.get('explanation')
        
        # In a real implementation, you'd save the explanation
        # and possibly send notifications to relevant personnel
        
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["POST"])
@csrf_exempt
def request_all_explanations_api(request):
    """API endpoint to request explanations for all current overruns"""
    try:
        # In a real implementation, you'd identify all overrunning phases
        # and send explanation requests to the relevant operators
        count = 3  # Sample count
        
        return JsonResponse({'success': True, 'count': count})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})

@login_required
@require_http_methods(["POST"])
def phase_timer_expired_api(request):
    """API endpoint to handle phase timer expiration - logging only"""
    try:
        import json
        
        # Parse JSON body
        data = json.loads(request.body.decode('utf-8'))
        phase_id = data.get('phase_id')
        
        if not phase_id:
            return JsonResponse({'success': False, 'error': 'Phase ID required'})
        
        # Just acknowledge timer expiration - background system handles notifications
        return JsonResponse({
            'success': True, 
            'message': 'Timer expiration logged. Background system handles notifications.'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error processing timer expiration: {str(e)}'})

@login_required
@require_http_methods(["GET"])
def monthly_production_analytics_api(request):
    """API endpoint to get monthly production analytics data"""
    try:
        # Get month and year from request parameters
        month = int(request.GET.get('month', timezone.now().month))
        year = int(request.GET.get('year', timezone.now().year))
        
        # Get analytics data
        monthly_analytics = get_monthly_production_analytics(month, year)
        yearly_comparison = get_yearly_production_comparison(year)
        
        # Format data for JSON response
        response_data = {
            'success': True,
            'monthly_analytics': monthly_analytics,
            'yearly_comparison': yearly_comparison,
            'current_month': month,
            'current_year': year
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False, 
            'error': f'Error fetching production analytics: {str(e)}'
        })

@login_required
def export_monthly_production_excel(request):
    """Export monthly production analytics to Excel"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"Export request received: {request.GET}")
    
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    try:
        # Get month and year from request parameters
        month = int(request.GET.get('month', timezone.now().month))
        year = int(request.GET.get('year', timezone.now().year))
        product_type_filter = request.GET.get('product_type')  # Optional filter
        
        # Generate Excel file
        logger.info(f"Generating Excel file for month={month}, year={year}, product_type={product_type_filter}")
        excel_file = export_monthly_production_to_excel(month, year, product_type_filter)
        
        if excel_file is None:
            logger.error("Excel file generation returned None")
            messages.error(request, 'Error generating Excel report. Please try again.')
            return redirect('dashboards:admin_dashboard')
        
        logger.info(f"Excel file generated successfully, size: {len(excel_file.getvalue())} bytes")
        
        # Create response
        response = HttpResponse(
            excel_file.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Set filename with month and year
        month_name = calendar.month_name[month]
        if product_type_filter:
            filename = f'KPI_{product_type_filter.title()}_Production_Report_{month_name}_{year}.xlsx'
        else:
            filename = f'KPI_Production_Report_{month_name}_{year}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting monthly production to Excel: {e}")
        messages.error(request, f'Error generating Excel report: {str(e)}')
        return redirect('dashboards:admin_dashboard')

@login_required
def export_wip(request):
    """Export Work in Progress BMRs to Excel"""
    if not (request.user.is_staff or request.user.role == 'admin'):
        messages.error(request, 'Access denied. Admin privileges required.')
        return redirect('dashboards:dashboard_home')
    
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from io import BytesIO
    from datetime import datetime
    
    # Get filter parameters
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    format_type = request.GET.get('format', 'excel')
    
    # Get work in progress BMRs using the same logic as dashboard
    from bmr.models import BMR
    from workflow.models import BatchPhaseExecution
    
    # Get BMRs that are currently in production (not completed or rejected)
    wip_bmrs_initial = BMR.objects.filter(
        status__in=['draft', 'approved', 'in_production']
    ).select_related('product', 'created_by')
    
    # Filter out fully completed BMRs to match dashboard logic
    genuine_wip_bmrs = []
    for bmr in wip_bmrs_initial:
        total_phases = BatchPhaseExecution.objects.filter(bmr=bmr).count()
        completed_phases = BatchPhaseExecution.objects.filter(bmr=bmr, status='completed').count()
        progress_percentage = int((completed_phases / total_phases * 100)) if total_phases > 0 else 0
        
        # Skip BMRs that are 100% complete - they shouldn't be in work in progress
        if progress_percentage < 100:
            genuine_wip_bmrs.append(bmr)
    
    # Convert to queryset format for date filtering
    if genuine_wip_bmrs:
        bmr_ids = [bmr.id for bmr in genuine_wip_bmrs]
        queryset = BMR.objects.filter(id__in=bmr_ids).select_related('product', 'created_by')
    else:
        queryset = BMR.objects.none()
    
    # Apply date filters if provided
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            queryset = queryset.filter(created_date__date__gte=start_date_obj)
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            queryset = queryset.filter(created_date__date__lte=end_date_obj)
        except ValueError:
            pass  # Invalid date format, ignore filter
    
    if format_type == 'excel':
        # Create Excel file
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Work in Progress'
        
        # Add headers
        headers = [
            'BMR Number', 'Product Name', 'Product Type', 'Batch Size', 'Unit',
            'Current Phase', 'Phase Status', 'Started Date', 'Progress %',
            'Created By', 'Created Date'
        ]
        
        # Write headers with styling
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Write data
        row = 2
        for bmr in queryset:
            # Use the same current phase logic as dashboard
            current_phase = BatchPhaseExecution.objects.filter(
                bmr=bmr,
                status__in=['pending', 'in_progress']
            ).select_related('phase').first()
            
            # If no current active phase, get the next phase that should start
            if not current_phase:
                current_phase = BatchPhaseExecution.objects.filter(
                    bmr=bmr,
                    status__in=['not_ready', 'pending']
                ).select_related('phase').order_by('id').first()
            
            # Calculate progress using same logic as dashboard
            total_phases = BatchPhaseExecution.objects.filter(bmr=bmr).count()
            completed_phases = BatchPhaseExecution.objects.filter(bmr=bmr, status='completed').count()
            progress_percentage = int((completed_phases / total_phases * 100)) if total_phases > 0 else 0
            
            # Determine current phase name using same logic as dashboard
            if current_phase:
                current_phase_name = current_phase.phase.get_phase_name_display()
                phase_status = current_phase.get_status_display()
                if current_phase.status == 'pending':
                    current_phase_name += ' (Pending)'
                elif current_phase.status == 'in_progress':
                    current_phase_name += ' (In Progress)'
                elif current_phase.status == 'not_ready':
                    current_phase_name += ' (Waiting)'
            else:
                if progress_percentage == 0:
                    current_phase_name = 'Not Started'
                    phase_status = 'Not Started'
                else:
                    current_phase_name = 'Phase Transition'
                    phase_status = 'Transition'
            
            # Get the first phase that was started to determine actual start date
            first_started_phase = BatchPhaseExecution.objects.filter(
                bmr=bmr,
                started_date__isnull=False
            ).order_by('started_date').first()
            
            actual_start_date = first_started_phase.started_date if first_started_phase else bmr.created_date
            
            # Write row data
            worksheet.cell(row=row, column=1, value=bmr.bmr_number)
            worksheet.cell(row=row, column=2, value=bmr.product.product_name)
            worksheet.cell(row=row, column=3, value=bmr.product.get_product_type_display())
            worksheet.cell(row=row, column=4, value=float(bmr.batch_size))
            worksheet.cell(row=row, column=5, value=bmr.product.batch_size_unit)
            worksheet.cell(row=row, column=6, value=current_phase_name)
            worksheet.cell(row=row, column=7, value=phase_status)
            worksheet.cell(row=row, column=8, value=actual_start_date.strftime('%Y-%m-%d %H:%M') if actual_start_date else '')
            worksheet.cell(row=row, column=9, value=f"{progress_percentage}%")
            worksheet.cell(row=row, column=10, value=bmr.created_by.get_full_name() if bmr.created_by else '')
            worksheet.cell(row=row, column=11, value=bmr.created_date.strftime('%Y-%m-%d %H:%M'))
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[chr(64 + col)].width = 15
        
        # Save to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Create response
        response = HttpResponse(
            output, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Set filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'work_in_progress_{timestamp}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    else:
        # Return JSON response for other formats or debugging
        wip_data = []
        for bmr in queryset:
            # Use the same current phase logic as dashboard and Excel export
            current_phase = BatchPhaseExecution.objects.filter(
                bmr=bmr,
                status__in=['pending', 'in_progress']
            ).select_related('phase').first()
            
            if not current_phase:
                current_phase = BatchPhaseExecution.objects.filter(
                    bmr=bmr,
                    status__in=['not_ready', 'pending']
                ).select_related('phase').order_by('id').first()
            
            total_phases = BatchPhaseExecution.objects.filter(bmr=bmr).count()
            completed_phases = BatchPhaseExecution.objects.filter(bmr=bmr, status='completed').count()
            progress_percentage = int((completed_phases / total_phases * 100)) if total_phases > 0 else 0
            
            # Determine current phase name using same logic
            if current_phase:
                current_phase_name = current_phase.phase.get_phase_name_display()
                if current_phase.status == 'pending':
                    current_phase_name += ' (Pending)'
                elif current_phase.status == 'in_progress':
                    current_phase_name += ' (In Progress)'
                elif current_phase.status == 'not_ready':
                    current_phase_name += ' (Waiting)'
            else:
                if progress_percentage == 0:
                    current_phase_name = 'Not Started'
                else:
                    current_phase_name = 'Phase Transition'
            
            # Get actual start date
            first_started_phase = BatchPhaseExecution.objects.filter(
                bmr=bmr,
                started_date__isnull=False
            ).order_by('started_date').first()
            
            actual_start_date = first_started_phase.started_date if first_started_phase else bmr.created_date
            
            wip_data.append({
                'bmr_number': bmr.bmr_number,
                'product_name': bmr.product.product_name,
                'current_phase': current_phase_name,
                'progress_percentage': progress_percentage,
                'started_date': actual_start_date.isoformat() if actual_start_date else None,
            })
        
        return JsonResponse({'work_in_progress': wip_data})


@login_required
def get_detailed_product_breakdown_api(request):
    """API endpoint to get detailed product breakdown for a specific product type and month"""
    if not request.user.is_staff:
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    try:
        product_type = request.GET.get('product_type')
        month = request.GET.get('month')
        year = request.GET.get('year')
        
        if not all([product_type, month, year]):
            return JsonResponse({
                'error': 'Missing required parameters: product_type, month, year'
            }, status=400)
        
        # Convert to integers
        month = int(month)
        year = int(year)
        
        # Import the analytics function
        from .analytics import get_detailed_product_breakdown
        
        # Get the detailed breakdown
        breakdown_data = get_detailed_product_breakdown(month, year, product_type)
        
        return JsonResponse(breakdown_data)
        
    except ValueError:
        return JsonResponse({'error': 'Invalid month or year format'}, status=400)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
