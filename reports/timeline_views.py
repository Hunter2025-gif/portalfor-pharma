from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Q
from bmr.models import BMR
from workflow.models import BatchPhaseExecution, ProductionPhase
from workflow.services import WorkflowService
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

@login_required
def timeline_list_view(request):
    """Enhanced timeline list view showing all BMRs with visual progress"""
    # Check if user is admin/staff - they see all BMRs
    is_admin = request.user.is_staff or request.user.is_superuser or request.user.role == 'admin'
    
    if is_admin:
        bmrs = BMR.objects.all().select_related('product', 'created_by').order_by('-created_date')
    else:
        # Operators only see BMRs they were involved in
        bmrs = BMR.objects.filter(
            Q(created_by=request.user) | Q(approved_by=request.user)
        ).select_related('product', 'created_by').order_by('-created_date')
    
    # Add progress information to each BMR
    bmr_progress = []
    stats = {'completed': 0, 'in_progress': 0, 'partially_complete': 0, 'not_started': 0}
    
    for bmr in bmrs:
        phases = BatchPhaseExecution.objects.filter(bmr=bmr)
        total_phases = phases.count()
        completed_phases = phases.filter(status='completed').count()
        in_progress_phases = phases.filter(status='in_progress').count()
        
        progress_percentage = (completed_phases / total_phases * 100) if total_phases > 0 else 0
        
        # Determine status
        if completed_phases == total_phases and total_phases > 0:
            status = 'completed'
            status_class = 'success'
        elif in_progress_phases > 0:
            status = 'in_progress'
            status_class = 'primary'
        elif completed_phases > 0:
            status = 'partially_complete'
            status_class = 'warning'
        else:
            status = 'not_started'
            status_class = 'secondary'
        
        # Update stats
        stats[status] += 1
        
        bmr_progress.append({
            'bmr': bmr,
            'total_phases': total_phases,
            'completed_phases': completed_phases,
            'progress_percentage': progress_percentage,
            'status': status,
            'status_class': status_class
        })
    
    context = {
        'bmr_progress': bmr_progress,
        'is_admin': is_admin,
        'total_bmrs': len(bmr_progress),
        'stats': stats,
    }
    
    return render(request, 'reports/timeline_list.html', context)

@login_required
def enhanced_timeline_view(request, bmr_id):
    """Enhanced timeline view with visual progress tracking"""
    bmr = get_object_or_404(BMR, id=bmr_id)
    
    # Get BMR request information if it exists
    from bmr.models import BMRRequest
    bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
    
    # Check if user has access to this BMR
    is_admin = request.user.is_staff or request.user.is_superuser or request.user.role == 'admin'
    
    if not is_admin:
        # Check user access
        user_bmrs = BMR.objects.filter(
            Q(created_by=request.user) | Q(approved_by=request.user)
        )
        user_phases = BatchPhaseExecution.objects.filter(
            Q(started_by=request.user) | Q(completed_by=request.user),
            bmr=bmr
        )
        
        if not (bmr in user_bmrs or user_phases.exists()):
            from django.contrib import messages
            messages.error(request, 'Access denied. You can only view BMRs you were involved in.')
            return redirect('reports:comments_report')
    
    # Get all phases for this BMR
    phase_executions = BatchPhaseExecution.objects.filter(bmr=bmr).select_related(
        'phase', 'started_by', 'completed_by'
    ).order_by('phase__phase_order')
    
    # Calculate overall progress
    total_phases = phase_executions.count()
    completed_phases = phase_executions.filter(status='completed').count()
    overall_progress = (completed_phases / total_phases * 100) if total_phases > 0 else 0
    
    # Find current phase
    current_phase = None
    next_phase = None
    
    for phase in phase_executions:
        if phase.status in ['pending', 'in_progress']:
            if not current_phase:
                current_phase = phase
            elif phase.status == 'in_progress':
                current_phase = phase
                break
    
    # Find next phase
    if current_phase:
        next_phases = phase_executions.filter(
            phase__phase_order__gt=current_phase.phase.phase_order,
            status='pending'
        ).order_by('phase__phase_order')
        if next_phases.exists():
            next_phase = next_phases.first()
    
    # Group phases into implementation stages (like the diagram)
    phase_groups = []
    
    # Development Phase (Early phases)
    development_phases = phase_executions.filter(
        phase__phase_name__in=[
            'material_dispensing', 'mixing', 'granulation', 'blending'
        ]
    )
    
    # Quality Control Phase (QC phases)
    qc_phases = phase_executions.filter(
        phase__phase_name__in=[
            'post_mixing_qc', 'post_blending_qc', 'post_compression_qc'
        ]
    )
    
    # Production Phase (Main production)
    production_phases = phase_executions.filter(
        phase__phase_name__in=[
            'compression', 'coating', 'drying', 'filling', 'tube_filling'
        ]
    )
    
    # Packaging Phase (Packaging operations)
    packaging_phases = phase_executions.filter(
        phase__phase_name__in=[
            'sorting', 'bulk_packing', 'packaging_material_release', 
            'blister_packing', 'secondary_packaging'
        ]
    )
    
    # Final Phase (Final operations)
    final_phases = phase_executions.filter(
        phase__phase_name__in=[
            'final_qa', 'finished_goods_store'
        ]
    )
    
    # Calculate progress for each group
    def calculate_group_progress(phases):
        if not phases.exists():
            return 0, 'not-ready'
        completed = phases.filter(status='completed').count()
        in_progress = phases.filter(status='in_progress').count()
        total = phases.count()
        
        if completed == total:
            return 100, 'completed'
        elif in_progress > 0 or completed > 0:
            return (completed / total * 100), 'in-progress'
        else:
            return 0, 'pending'
    
    dev_progress, dev_status = calculate_group_progress(development_phases)
    qc_progress, qc_status = calculate_group_progress(qc_phases)
    prod_progress, prod_status = calculate_group_progress(production_phases)
    pack_progress, pack_status = calculate_group_progress(packaging_phases)
    final_progress, final_status = calculate_group_progress(final_phases)
    
    phase_groups = [
        {
            'name': 'Planning & Setup',
            'icon': 'fas fa-lightbulb',
            'color': 'warning',
            'phases': development_phases,
            'progress': dev_progress,
            'status': dev_status,
            'description': 'Material preparation and initial processing'
        },
        {
            'name': 'Quality Control',
            'icon': 'fas fa-microscope',
            'color': 'info',
            'phases': qc_phases,
            'progress': qc_progress,
            'status': qc_status,
            'description': 'Quality testing and validation'
        },
        {
            'name': 'Production',
            'icon': 'fas fa-cogs',
            'color': 'primary',
            'phases': production_phases,
            'progress': prod_progress,
            'status': prod_status,
            'description': 'Core manufacturing operations'
        },
        {
            'name': 'Packaging',
            'icon': 'fas fa-boxes',
            'color': 'success',
            'phases': packaging_phases,
            'progress': pack_progress,
            'status': pack_status,
            'description': 'Packaging and material handling'
        },
        {
            'name': 'Go Live!',
            'icon': 'fas fa-rocket',
            'color': 'dark',
            'phases': final_phases,
            'progress': final_progress,
            'status': final_status,
            'description': 'Final approval and storage'
        }
    ]
    
    # Calculate total production time
    total_production_time = None
    total_production_hours = 0
    production_status = "In Progress"
    
    # Calculate duration for each phase and add to phase objects
    for phase in phase_executions:
        if phase.started_date and phase.completed_date:
            phase_duration = phase.completed_date - phase.started_date
            # Add duration in hours as a dynamic attribute to the phase object
            phase.calculated_duration_hours = round(phase_duration.total_seconds() / 3600, 2)
        else:
            phase.calculated_duration_hours = None
    
    # Calculate total production time from BMR request (if available) to FGS completion
    start_date = bmr_request.request_date if bmr_request else bmr.created_date
    fgs_phase = phase_executions.filter(phase__phase_name='finished_goods_store', status='completed').first()
    
    # If we have both start and end dates, calculate total production time
    if start_date and fgs_phase and fgs_phase.completed_date:
        total_duration = fgs_phase.completed_date - start_date
        total_production_hours = total_duration.total_seconds() / 3600
    else:
        # If no FGS completion, calculate time from start to now for in-progress tracking
        if start_date:
            from django.utils import timezone
            total_duration = timezone.now() - start_date
            total_production_hours = total_duration.total_seconds() / 3600
        else:
            total_production_hours = 0
    
    # If all phases are completed, mark as completed
    if completed_phases == total_phases and total_phases > 0:
        production_status = "Completed"
        if total_production_hours > 0:
            days = int(total_production_hours // 24)
            hours = int(total_production_hours % 24)
            minutes = int((total_production_hours % 1) * 60)
            if days > 0:
                total_production_time = f"{days}d {hours}h"
            elif hours > 0:
                total_production_time = f"{hours}h {minutes}m"
            else:
                total_production_time = f"{minutes}m"
        else:
            total_production_time = "Completed"
    else:
        # Calculate time so far for in-progress batches
        if total_production_hours > 0:
            days = int(total_production_hours // 24)
            hours = int(total_production_hours % 24)
            minutes = int((total_production_hours % 1) * 60)
            if days > 0:
                total_production_time = f"{days}d {hours}h (So Far)"
            elif hours > 0:
                total_production_time = f"{hours}h {minutes}m (So Far)"
            else:
                total_production_time = f"{minutes}m (So Far)"
        else:
            total_production_time = "In Progress"

    # Add formatted durations for display
    for phase in phase_executions:
        # First try the calculated duration attribute
        if hasattr(phase, 'calculated_duration_hours') and phase.calculated_duration_hours is not None:
            hours = int(phase.calculated_duration_hours)
            minutes = int((phase.calculated_duration_hours - hours) * 60)
            if hours > 0:
                phase.duration = f"{hours}h {minutes}m"
            else:
                phase.duration = f"{minutes}m"
        # Fallback to the property if it exists and has a value
        elif phase.duration_hours is not None:
            hours = int(phase.duration_hours)
            minutes = int((phase.duration_hours - hours) * 60)
            if hours > 0:
                phase.duration = f"{hours}h {minutes}m"
            else:
                phase.duration = f"{minutes}m"
        else:
            phase.duration = "--"
    
    context = {
        'bmr': bmr,
        'bmr_request': bmr_request,
        'phase_executions': phase_executions,
        'overall_progress': overall_progress,
        'current_phase': current_phase,
        'next_phase': next_phase,
        'phase_groups': phase_groups,
        'total_phases': total_phases,
        'completed_phases': completed_phases,
        'remaining_phases': total_phases - completed_phases,
        'total_production_time': total_production_time,
        'total_production_hours': total_production_hours,
        'production_status': production_status,
        'is_admin': is_admin
    }
    
    return render(request, 'reports/enhanced_timeline.html', context)

@login_required
def export_timeline_csv(request):
    """Export BMR timeline data to CSV"""
    # Check user permissions
    is_admin = request.user.is_staff or request.user.is_superuser or request.user.role == 'admin'
    
    if is_admin:
        bmrs = BMR.objects.all().select_related('product', 'created_by').order_by('-created_date')
    else:
        bmrs = BMR.objects.filter(
            Q(created_by=request.user) | Q(approved_by=request.user)
        ).select_related('product', 'created_by').order_by('-created_date')
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bmr_timeline_export.csv"'
    
    writer = csv.writer(response)
    
    # Write header with BMR request fields
    writer.writerow([
        'BMR Number',
        'Product Name', 
        'Product Type',
        'BMR Requested Date',
        'BMR Requested By',
        'BMR Request Status',
        'BMR Created Date',
        'Created By',
        'BMR Status',
        'Total Phases',
        'Completed Phases',
        'In Progress Phases',
        'Progress %',
        'Current Phase',
        'Current Phase Status',
        'Last Updated',
        'Time in Current Phase (hours)',
        'Total Time Since Request (hours)'
    ])
    
    # Write data for each BMR
    for bmr in bmrs:
        # Get BMR request information
        from bmr.models import BMRRequest
        bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
        
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase')
        total_phases = phases.count()
        completed_phases = phases.filter(status='completed').count()
        in_progress_phases = phases.filter(status='in_progress').count()
        
        progress_percentage = (completed_phases / total_phases * 100) if total_phases > 0 else 0
        
        # Find current phase (first pending or in_progress phase)
        current_phase = phases.filter(status__in=['pending', 'in_progress']).order_by('phase__phase_order').first()
        current_phase_name = current_phase.phase.phase_name if current_phase else 'Completed'
        current_phase_status = current_phase.status if current_phase else 'completed'
        
        # Calculate time in current phase
        time_in_phase = ''
        if current_phase and current_phase.started_date:
            from django.utils import timezone
            time_diff = timezone.now() - current_phase.started_date
            hours = int(time_diff.total_seconds() / 3600)
            time_in_phase = f"{hours}"
        
        # Calculate total time since request
        total_time_since_request = ''
        if bmr_request and bmr_request.request_date:
            from django.utils import timezone
            time_diff = timezone.now() - bmr_request.request_date
            hours = int(time_diff.total_seconds() / 3600)
            total_time_since_request = f"{hours}"
        
        # Get last updated date (most recent phase activity)
        last_phase = phases.exclude(started_date__isnull=True).order_by('-started_date').first()
        last_updated = last_phase.started_date if last_phase else bmr.created_date
        
        writer.writerow([
            bmr.batch_number,
            bmr.product.product_name,
            bmr.product.product_type,
            bmr_request.request_date.strftime('%Y-%m-%d %H:%M') if bmr_request and bmr_request.request_date else 'N/A',
            bmr_request.requested_by.get_full_name() if bmr_request and bmr_request.requested_by else 'N/A',
            bmr_request.get_status_display() if bmr_request else 'N/A',
            bmr.created_date.strftime('%Y-%m-%d %H:%M'),
            bmr.created_by.get_full_name() if bmr.created_by else 'Unknown',
            bmr.status,
            total_phases,
            completed_phases,
            in_progress_phases,
            round(progress_percentage, 1),
            current_phase_name,
            current_phase_status,
            last_updated.strftime('%Y-%m-%d %H:%M') if last_updated else '',
            time_in_phase,
            total_time_since_request
        ])
    
    return response

@login_required
def export_timeline_excel(request):
    """Export comprehensive BMR timeline data to Excel matching the original format"""
    from django.utils import timezone
    from datetime import datetime
    
    # Check user permissions
    is_admin = request.user.is_staff or request.user.is_superuser or request.user.role == 'admin'
    
    if is_admin:
        bmrs = BMR.objects.all().select_related('product', 'created_by').order_by('-created_date')
    else:
        bmrs = BMR.objects.filter(
            Q(created_by=request.user) | Q(approved_by=request.user)
        ).select_related('product', 'created_by').order_by('-created_date')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=14)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    # Remove default sheet and create Production Summary
    wb.remove(wb.active)
    ws_summary = wb.create_sheet("Production Summary")
    
    # Add title and report info
    ws_summary.merge_cells('A1:L1')
    title_cell = ws_summary['A1']
    title_cell.value = "Kampala Pharmaceutical Industries - BMR Production Timeline Summary"
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    
    ws_summary.merge_cells('A2:L2') 
    report_cell = ws_summary['A2']
    report_cell.value = f"Report Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    report_cell.alignment = center_alignment
    
    # Headers for summary (row 4)
    summary_headers = [
        'Batch Number', 'Product Name', 'Product Type', 'Request Date',
        'Created Date', 'Current Status', 'Current Phase', 
        'Total Duration (Hours)', 'Completed', 'Bottleneck Phase'
    ]
    
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Fill summary data
    summary_row = 5
    for bmr in bmrs:
        # Get BMR request information
        from bmr.models import BMRRequest
        bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
        
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase')
        
        # Calculate total duration correctly (start to end, not sum of phases)
        total_duration = 0
        first_started_phase = phases.filter(started_date__isnull=False).order_by('started_date').first()
        last_completed_phase = phases.filter(completed_date__isnull=False).order_by('-completed_date').first()
        
        if first_started_phase and last_completed_phase:
            total_duration = (last_completed_phase.completed_date - first_started_phase.started_date).total_seconds() / 3600
        
        # Find current phase
        current_phase = phases.filter(status__in=['pending', 'in_progress']).order_by('phase__phase_order').first()
        current_phase_name = current_phase.phase.phase_name.replace('_', ' ').title() if current_phase else 'Completed'
        
        # Determine completion status
        completed_phases = phases.filter(status='completed').count()
        total_phases = phases.count()
        is_completed = "Yes" if completed_phases == total_phases else "No"
        
        # Find bottleneck (longest phase)
        bottleneck_phase = "Bmr Creation"  # Default
        max_duration = 0
        for phase in phases:
            if phase.started_date and phase.completed_date:
                phase_duration = (phase.completed_date - phase.started_date).total_seconds() / 3600
                if phase_duration > max_duration:
                    max_duration = phase_duration
                    bottleneck_phase = phase.phase.phase_name.replace('_', ' ').title()
        
        created_date_naive = bmr.created_date.replace(tzinfo=None) if bmr.created_date else None
        request_date_naive = bmr_request.request_date.replace(tzinfo=None) if bmr_request and bmr_request.request_date else None
        
        row_data = [
            bmr.batch_number,
            bmr.product.product_name,
            bmr.product.product_type.title(),
            request_date_naive if request_date_naive else 'N/A',
            created_date_naive,
            bmr.status.replace('_', ' ').title(),
            current_phase_name,
            "In Progress" if total_duration == 0 else f"{total_duration:.1f}",
            is_completed,
            bottleneck_phase
        ]
        
        for col, value in enumerate(row_data, 1):
            ws_summary.cell(row=summary_row, column=col, value=value)
        
        summary_row += 1
    
    # Now create detailed timeline sheets for each BMR
    for bmr in bmrs:
        # Create sheet for each BMR
        sheet_name = f"BMR-{bmr.batch_number}"
        ws_detail = wb.create_sheet(sheet_name)
        
        # Title for detailed sheet
        ws_detail.merge_cells('A1:Q1')
        detail_title = ws_detail['A1']
        detail_title.value = f"Detailed Timeline for BMR {bmr.batch_number} - {bmr.product.product_name}"
        detail_title.font = title_font
        detail_title.alignment = center_alignment
        
        # Product info (row 2)
        ws_detail['A2'] = f"Product Type: {bmr.product.product_type} | Created: {bmr.created_date.strftime('%Y-%m-%d %H:%M:%S') if bmr.created_date else 'N/A'}"
        
        # Calculate total production time for this BMR correctly
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related('phase')
        total_production_hours = 0
        completed_phases_count = 0
        total_phases_count = phases.count()
        
        # Debug: Check each phase status
        completed_phases = phases.filter(status='completed')
        completed_phases_count = completed_phases.count()
        
        # Calculate correct total production time (start to end, not sum of phases)
        first_started_phase = phases.filter(started_date__isnull=False).order_by('started_date').first()
        last_completed_phase = phases.filter(completed_date__isnull=False).order_by('-completed_date').first()
        
        if first_started_phase and last_completed_phase:
            total_duration = last_completed_phase.completed_date - first_started_phase.started_date
            total_production_hours = total_duration.total_seconds() / 3600
        elif first_started_phase:
            # For in-progress BMRs, calculate from start to now
            from django.utils import timezone
            total_duration = timezone.now() - first_started_phase.started_date
            total_production_hours = total_duration.total_seconds() / 3600
        
        # Format production time display
        if completed_phases_count == total_phases_count and total_phases_count > 0:
            # All phases completed
            if total_production_hours > 0:
                days = int(total_production_hours // 24)
                hours = int(total_production_hours % 24)
                minutes = int((total_production_hours % 1) * 60)
                if days > 0:
                    production_time_display = f"Completed in {days}d {hours}h"
                elif hours > 0:
                    production_time_display = f"Completed in {hours}h {minutes}m"
                else:
                    production_time_display = f"Completed in {minutes}m"
            else:
                production_time_display = "Completed"
        else:
            # In progress
            if total_production_hours > 0:
                days = int(total_production_hours // 24)
                hours = int(total_production_hours % 24)
                minutes = int((total_production_hours % 1) * 60)
                if days > 0:
                    production_time_display = f"In Progress ({days}d {hours}h so far)"
                elif hours > 0:
                    production_time_display = f"In Progress ({hours}h {minutes}m so far)"
                else:
                    production_time_display = f"In Progress ({minutes}m so far)"
            else:
                production_time_display = "In Progress"
        
        ws_detail['A3'] = f"Total Production Time: {production_time_display}"
        
        # Detailed headers (row 5) - matching the Excel format exactly
        detail_headers = [
            'Phase Name', 'Status', 'Started Date', 'Started By', 'Completed Date', 
            'Complete', 'Duration (Hours)', 'Comments', 'Machine Used',
            'Breakdown Occurred', 'Breakdown Duration', 'Breakdown down',
            'Breakdown End', 'Changeover Occurred', 'Changeover Duration (Min)',
            'Changeover Start', 'Changeover End'
        ]
        
        for col, header in enumerate(detail_headers, 1):
            cell = ws_detail.cell(row=5, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
        
        # Get phases for this BMR
        phases = BatchPhaseExecution.objects.filter(bmr=bmr).select_related(
            'phase', 'started_by', 'completed_by', 'machine_used'
        ).order_by('phase__phase_order')
        
        # Get BMR request information
        from bmr.models import BMRRequest
        bmr_request = BMRRequest.objects.filter(bmr=bmr).first()
        
        detail_row = 6
        
        # Add BMR Request as first row if it exists
        if bmr_request:
            request_date_naive = bmr_request.request_date.replace(tzinfo=None) if bmr_request.request_date else None
            # Calculate time from request to BMR creation
            request_duration = ''
            if bmr_request.request_date and bmr.created_date:
                time_diff = bmr.created_date - bmr_request.request_date
                duration_hours = time_diff.total_seconds() / 3600
                if duration_hours >= 1:
                    request_duration = round(duration_hours, 1)
                else:
                    minutes = round(duration_hours * 60, 1)
                    request_duration = f"{minutes}m"
            
            bmr_created_date_naive = bmr.created_date.replace(tzinfo=None) if bmr.created_date else None
            
            request_data = [
                'BMR Request Submitted',  # Phase Name
                bmr_request.get_status_display(),  # Status
                request_date_naive,  # Started Date
                bmr_request.requested_by.get_full_name() if bmr_request.requested_by else '',  # Started By
                bmr_created_date_naive if bmr_request.status == 'completed' else '',  # Completed Date
                bmr_created_date_naive if bmr_request.status == 'completed' else 'Not Completed',  # Complete
                request_duration,  # Duration (Hours)
                f"Priority: {bmr_request.get_priority_display()}, Reason: {bmr_request.reason}" if bmr_request.reason else f"Priority: {bmr_request.get_priority_display()}",  # Comments
                '',  # Machine Used
                'No',  # Breakdown Occurred
                '',  # Breakdown Duration
                '',  # Breakdown Start
                '',  # Breakdown End
                'No',  # Changeover Occurred
                '',  # Changeover Duration
                '',  # Changeover Start
                ''   # Changeover End
            ]
            
            for col, value in enumerate(request_data, 1):
                cell = ws_detail.cell(row=detail_row, column=col, value=value)
                # Apply special formatting for BMR request row
                if col == 2:  # Status column
                    if bmr_request.status == 'completed':
                        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                    elif bmr_request.status == 'pending':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            detail_row += 1
        
        for phase in phases:
            # Calculate duration in hours
            duration = ''
            breakdown_duration = ''
            changeover_duration = ''
            breakdown_occurred = 'No'
            changeover_occurred = 'No'
            
            # Main phase duration - FIXED to ensure it shows up
            if phase.started_date and phase.completed_date:
                time_diff = phase.completed_date - phase.started_date
                duration_hours = time_diff.total_seconds() / 3600
                # Ensure duration is always shown with proper formatting
                if duration_hours >= 1:
                    duration = round(duration_hours, 1)
                else:
                    # Show minutes if less than an hour
                    minutes = round(duration_hours * 60, 1)
                    duration = f"{minutes}m"
            elif phase.started_date:
                # Phase is in progress - calculate time so far
                from django.utils import timezone
                time_diff = timezone.now() - phase.started_date
                duration_hours = time_diff.total_seconds() / 3600
                if duration_hours >= 1:
                    duration = f"{round(duration_hours, 1)} (ongoing)"
                else:
                    minutes = round(duration_hours * 60, 1)
                    duration = f"{minutes}m (ongoing)"
            else:
                # Phase not started
                duration = "Not Started"
            
            # Breakdown duration (if exists)
            if hasattr(phase, 'breakdown_start_time') and hasattr(phase, 'breakdown_end_time'):
                if phase.breakdown_start_time and phase.breakdown_end_time:
                    breakdown_occurred = 'Yes'
                    breakdown_diff = phase.breakdown_end_time - phase.breakdown_start_time
                    breakdown_duration = round(breakdown_diff.total_seconds() / 3600, 1)
            
            # Changeover duration (if exists)
            if hasattr(phase, 'changeover_start_time') and hasattr(phase, 'changeover_end_time'):
                if phase.changeover_start_time and phase.changeover_end_time:
                    changeover_occurred = 'Yes'
                    changeover_diff = phase.changeover_end_time - phase.changeover_start_time
                    changeover_duration = round(changeover_diff.total_seconds() / 60, 1)  # Minutes for changeover
            
            # Convert dates to naive
            started_date_naive = phase.started_date.replace(tzinfo=None) if phase.started_date else None
            completed_date_naive = phase.completed_date.replace(tzinfo=None) if phase.completed_date else None
            breakdown_start_naive = phase.breakdown_start_time.replace(tzinfo=None) if hasattr(phase, 'breakdown_start_time') and phase.breakdown_start_time else None
            breakdown_end_naive = phase.breakdown_end_time.replace(tzinfo=None) if hasattr(phase, 'breakdown_end_time') and phase.breakdown_end_time else None
            changeover_start_naive = phase.changeover_start_time.replace(tzinfo=None) if hasattr(phase, 'changeover_start_time') and phase.changeover_start_time else None
            changeover_end_naive = phase.changeover_end_time.replace(tzinfo=None) if hasattr(phase, 'changeover_end_time') and phase.changeover_end_time else None
            
            # Status mapping
            status_map = {
                'completed': 'Completed',
                'in_progress': 'Not Completed',
                'pending': 'Not Ready',
                'failed': 'Not Completed'
            }
            
            phase_data = [
                phase.phase.phase_name.replace('_', ' ').title(),
                status_map.get(phase.status, phase.status),
                "Not Started" if not started_date_naive else started_date_naive,
                phase.started_by.get_full_name() if phase.started_by else '',
                "Not Completed" if not completed_date_naive else completed_date_naive,
                "Not Completed" if phase.status != 'completed' else completed_date_naive,
                duration if duration else '',  # Duration (Hours)
                phase.operator_comments or '',  # Comments
                phase.machine_used.name if phase.machine_used else '',  # Machine Used
                breakdown_occurred,  # Breakdown Occurred
                breakdown_duration if breakdown_duration else '',  # Breakdown Duration
                breakdown_start_naive if breakdown_start_naive else '',  # Breakdown Start
                breakdown_end_naive if breakdown_end_naive else '',  # Breakdown End
                changeover_occurred,  # Changeover Occurred
                changeover_duration if changeover_duration else '',  # Changeover Duration (Min)
                changeover_start_naive if changeover_start_naive else '',  # Changeover Start
                changeover_end_naive if changeover_end_naive else ''  # Changeover End
            ]
            
            for col, value in enumerate(phase_data, 1):
                cell = ws_detail.cell(row=detail_row, column=col, value=value)
                # Apply conditional formatting based on status
                if col == 2:  # Status column
                    if phase.status == 'completed':
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif phase.status == 'in_progress':
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif phase.status == 'pending':
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            detail_row += 1
    
    # Auto-adjust column widths for all sheets
    for sheet in wb.worksheets:
        for column_letter in sheet.column_dimensions:
            max_length = 0
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.column_letter == column_letter and not hasattr(cell, '_merge_start'):
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
            # Set minimum width and maximum width
            adjusted_width = min(max(max_length + 2, 10), 30)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        # Also adjust any columns that weren't in column_dimensions
        from openpyxl.utils import get_column_letter
        for col_num in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(col_num)
            if column_letter not in sheet.column_dimensions:
                max_length = 0
                for row in sheet.iter_rows(min_col=col_num, max_col=col_num):
                    for cell in row:
                        if not hasattr(cell, '_merge_start'):
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                adjusted_width = min(max(max_length + 2, 10), 30)
                sheet.column_dimensions[column_letter].width = adjusted_width
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bmr_timeline_{timestamp}.xlsx"
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    return response
